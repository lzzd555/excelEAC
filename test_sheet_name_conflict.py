"""
测试 sheet 名称冲突的场景
"""
import pandas as pd
import openpyxl
import os

# 清理旧文件
for f in ['sales_data.xlsx', 'cost_data.xlsx', 'template_sheet_conflict.xlsx', 'test_sheet_conflict_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建数据源文件（sheet 名称都是 DataSheet）
def create_data_files():
    """创建数据源文件（相同 sheet 名称）"""
    # 数据源 1
    data1 = pd.DataFrame({'Sales': [100, 200, 300]})
    with pd.ExcelWriter('sales_data.xlsx', engine='openpyxl') as writer:
        data1.to_excel(writer, sheet_name='DataSheet', index=False)
    print("✓ 创建 sales_data.xlsx (DataSheet)")

    # 数据源 2
    data2 = pd.DataFrame({'Cost': [50, 100, 150]})
    with pd.ExcelWriter('cost_data.xlsx', engine='openpyxl') as writer:
        data2.to_excel(writer, sheet_name='DataSheet', index=False)
    print("✓ 创建 cost_data.xlsx (DataSheet)")


# 创建引用第二个数据源的模板
def create_template():
    """创建引用第二个数据源的模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'

    # 第一行：列名
    headers = ['Sales', 'Cost', 'ExternalRef', 'Total']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # 第二行：公式
    # ExternalRef: 引用第二个数据源 (应该是 cost_data.xlsx 的 DataSheet!A1)
    # 因为两个数据源的 sheet 名称都是 'DataSheet'，所以需要使用别名或索引
    # 这里使用索引格式 [2]
    ws.cell(row=2, column=3).value = "=[2]DataSheet!A1"

    # Total: =Sales + ExternalRef
    ws.cell(row=2, column=4).value = "=A2+C2"

    # 保存文件
    wb.save('template_sheet_conflict.xlsx')
    wb.close()

    print("✓ 创建模板文件: template_sheet_conflict.xlsx")
    print("  ExternalRef 公式: =[2]DataSheet!A1 (引用第二个数据源)")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：Sheet 名称冲突的外部引用映射")
    print("="*70 + "\n")

    # 数据源配置 - 顺序很重要
    data_sources = [
        {
            'file_path': 'sales_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [{'source': 'Sales', 'target': 'Sales'}],
            'alias': 'sales'  # 使用别名区分
        },
        {
            'file_path': 'cost_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [{'source': 'Cost', 'target': 'Cost'}],
            'alias': 'cost'   # 使用别名区分
        }
    ]

    # 公式列
    formula_columns = ['ExternalRef', 'Total']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_sheet_conflict.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_sheet_conflict_output.xlsx',
            use_external_refs=True
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_sheet_conflict_output.xlsx")

        # 验证输出文件中的公式
        print("\n正在验证输出文件中的公式...")
        wb = openpyxl.load_workbook('test_sheet_conflict_output.xlsx', data_only=False)
        ws = wb.active

        # 读取第二行的公式
        external_ref_formula = ws.cell(row=2, column=3).value
        total_formula = ws.cell(row=2, column=4).value

        print(f"\n公式验证结果:")
        print(f"  ExternalRef: {external_ref_formula}")
        print(f"  Total: {total_formula}")

        wb.close()

        # 检查引用是否正确
        print("\n结果分析:")
        print(f"  模板中的公式: =[2]DataSheet!A1")
        print(f"  期望输出公式: 应该引用 cost_data.xlsx 的 DataSheet (第二个数据源)")
        print(f"  实际输出公式: {external_ref_formula}")

        if 'cost_data.xlsx' in str(external_ref_formula) or 'Sales' in str(external_ref_formula):
            if 'cost_data.xlsx' in str(external_ref_formula):
                print("\n✓ 正确：公式引用了正确的数据源 (cost_data.xlsx)")
            else:
                print("\n✗ 错误：公式引用了错误的数据源 (应该是 cost_data.xlsx，但引用了 sales_data.xlsx)")
        else:
            print("\n⚠ 无法确定：公式可能使用了本地引用")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 创建测试文件
    create_data_files()
    create_template()

    # 运行测试
    run_test()
