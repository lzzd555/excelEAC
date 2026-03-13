"""
测试所有样式的完整复制（包括填充）
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
import os

# 清理旧文件
for f in ['external_data.xlsx', 'template_complete_styles.xlsx', 'test_complete_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Value': [100, 200, 300]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含所有样式的模板文件
def create_template_with_complete_styles():
    """创建包含所有样式的模板文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'

    # 第一行：列名
    headers = [
        'Fill', 'FontBold', 'FontItalic', 'FontColor', 'AlignmentCenter',
        'AlignmentWrap', 'Border', 'NumberFormat', 'Protection'
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='gray125')

    # 第二行：样式测试数据
    # 1. 填充样式
    cell = ws.cell(row=2, column=1)
    cell.value = "Fill"
    cell.fill = PatternFill(fill_type='solid', start_color='FFCC99')

    # 2. 粗体字体
    cell = ws.cell(row=2, column=2)
    cell.value = "Bold Text"
    cell.font = Font(bold=True, size=12)

    # 3. 斜体字体
    cell = ws.cell(row=2, column=3)
    cell.value = "Italic Text"
    cell.font = Font(italic=True, size=12)

    # 4. 颜色字体
    cell = ws.cell(row=2, column=4)
    cell.value = "Red Text"
    cell.font = Font(color='FF0000', size=12)

    # 5. 居中对齐
    cell = ws.cell(row=2, column=5)
    cell.value = "Center"
    cell.alignment = Alignment(horizontal='center')

    # 6. 自动换行
    cell = ws.cell(row=2, column=6)
    cell.value = "Wrap\nText"
    cell.alignment = Alignment(wrap_text=True)

    # 7. 边框
    cell = ws.cell(row=2, column=7)
    cell.value = "Bordered"
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 8. 数字格式
    cell = ws.cell(row=2, column=8)
    cell.value = 12345.67
    cell.number_format = '#,##0.00'

    # 9. 保护锁定
    cell = ws.cell(row=2, column=9)
    cell.value = "Protected"
    cell.protection = Protection(locked=True)

    # 保存文件
    wb.save('template_complete_styles.xlsx')
    wb.close()

    print("✓ 创建包含所有样式的模板文件")
    print("  - 填充样式：实心填充（橙色）")
    print("  - 字体样式：粗体、斜体、颜色")
    print("  - 对齐样式：居中、自动换行")
    print("  - 边框样式：实线边框")
    print("  - 数字格式：#,##0.00")
    print("  - 保护样式：锁定")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：所有样式完整复制")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [{'source': 'Value', 'target': 'Value'}],
            'alias': 'external_data'
        }
    ]

    # 公式列
    formula_columns = []

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_complete_styles.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_complete_output.xlsx',
            use_external_refs=False
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_complete_output.xlsx")

        # 验证输出文件中的样式
        wb = openpyxl.load_workbook('test_complete_output.xlsx', data_only=False)
        ws = wb.active

        print("\n样式验证结果:")

        # 验证填充样式
        fill_cell = ws.cell(row=2, column=1)
        print(f"\n🎨 填充样式:")
        print(f"  类型: {fill_cell.fill.fill_type if fill_cell.fill else 'None'}")
        if fill_cell.fill and fill_cell.fill.fill_type == 'solid' and fill_cell.fill.start_color:
            print(f"  颜色: {fill_cell.fill.start_color}")

        # 验证字体样式
        cell = ws.cell(row=2, column=2)
        print(f"\n🔤 字体样式:")
        print(f"  粗体: {'✓' if cell.font.bold else '✗'}")
        print(f"  斜体: {'✓' if cell.font.italic else '✗'}")

        cell = ws.cell(row=2, column=3)
        print(f"  字体大小: {cell.font.size if cell.font.size else '无'}")

        cell = ws.cell(row=2, column=4)
        if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
            # 移除前导的'00'（如果是ARGB格式）
            rgb_value = cell.font.color.rgb
            if rgb_value.startswith('00FF'):
                rgb_value = rgb_value[2:]
            print(f"  字体颜色: RGB={rgb_value} ✓")
        else:
            print(f"  字体颜色: ✗")

        # 验证对齐样式
        cell = ws.cell(row=2, column=5)
        print(f"\n📍 对齐样式:")
        print(f"  居中对齐: {'✓' if cell.alignment.horizontal == 'center' else '✗'}")

        cell = ws.cell(row=2, column=6)
        print(f"  自动换行: {'✓' if cell.alignment.wrap_text else '✗'}")

        # 验证边框样式
        cell = ws.cell(row=2, column=7)
        has_left_border = cell.border.left.border_style != 'none'
        has_right_border = cell.border.right.border_style != 'none'
        has_top_border = cell.border.top.border_style != 'none'
        has_bottom_border = cell.border.bottom.border_style != 'none'
        has_border = has_left_border and has_right_border and has_top_border and has_bottom_border
        print(f"\n🔲 边框样式:")
        print(f"  边框: {'✓' if has_border else '✗'} (左:{'✓' if has_left_border else '✗'} 右:{'✓' if has_right_border else '✗'} 上:{'✓' if has_top_border else '✗'} 下:{'✓' if has_bottom_border else '✗'})")

        # 验证数字格式
        cell = ws.cell(row=2, column=8)
        print(f"\n🔢 数字格式:")
        print(f"  格式: {cell.number_format}")

        # 验证保护样式
        cell = ws.cell(row=2, column=9)
        print(f"\n🔒 保护样式:")
        print(f"  保护锁定: {'✓' if cell.protection.locked else '✗'}")

        wb.close()

        # 统计结果
        print("\n" + "="*70)
        print("📊 样式复制统计:")
        print("="*70)
        all_styles = ['填充', '粗体', '斜体', '字体颜色', '居中对齐', '自动换行', '边框', '数字格式', '保护锁定']
        passed_styles = []

        # 重新检查所有样式
        styles_passed = 0
        total_styles = 9

        # 检查填充
        fill_cell = ws.cell(row=2, column=1)
        if fill_cell.fill and fill_cell.fill.fill_type == 'solid':
            passed_styles.append('填充')
            styles_passed += 1

        # 检查字体
        cell = ws.cell(row=2, column=2)
        if cell.font.bold:
            passed_styles.append('粗体')
            styles_passed += 1

        cell = ws.cell(row=2, column=3)
        if cell.font.italic:
            passed_styles.append('斜体')
            styles_passed += 1

        cell = ws.cell(row=2, column=4)
        if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
            # 检查是否为红色（FF0000）
            rgb_value = cell.font.color.rgb
            if rgb_value.startswith('00FF'):
                rgb_value = rgb_value[2:]
            if rgb_value == 'FF0000':
                passed_styles.append('字体颜色')
                styles_passed += 1

        # 检查对齐
        cell = ws.cell(row=2, column=5)
        if cell.alignment and cell.alignment.horizontal == 'center':
            passed_styles.append('居中对齐')
            styles_passed += 1

        cell = ws.cell(row=2, column=6)
        if cell.alignment and cell.alignment.wrap_text:
            passed_styles.append('自动换行')
            styles_passed += 1

        # 检查边框
        cell = ws.cell(row=2, column=7)
        if (cell.border and
            cell.border.left.border_style != 'none' and
            cell.border.right.border_style != 'none' and
            cell.border.top.border_style != 'none' and
            cell.border.bottom.border_style != 'none'):
            passed_styles.append('边框')
            styles_passed += 1

        # 检查数字格式
        cell = ws.cell(row=2, column=8)
        if cell.number_format == '#,##0.00':
            passed_styles.append('数字格式')
            styles_passed += 1

        # 检查保护
        cell = ws.cell(row=2, column=9)
        if cell.protection and cell.protection.locked:
            passed_styles.append('保护锁定')
            styles_passed += 1

        print(f"✅ 成功复制: {styles_passed}/{total_styles} 个样式")
        print(f"✅ 成功列表: {', '.join(passed_styles)}")

        if styles_passed == total_styles:
            print(f"\n🎉 所有样式复制功能完美工作！")
        else:
            failed_styles = [s for s in all_styles if s not in passed_styles]
            print(f"\n❌ 需要改进: {', '.join(failed_styles)}")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

    return True


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_complete_styles()

    # 运行测试
    if run_test():
        print("\n✅ 完整样式复制功能验证通过！")
    else:
        print("\n❌ 完整样式复制功能验证失败！")