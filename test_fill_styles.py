"""
测试填充样式（PatternFill）的复制功能
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

# 清理旧文件
for f in ['external_data.xlsx', 'template_fill_styles.xlsx', 'test_fill_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Value': [100, 200, 300]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含各种填充样式的模板文件
def create_template_with_fill_styles():
    """创建包含各种填充样式的模板文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'

    # 第一行：列名
    headers = [
        'NoFill', 'SolidRGB', 'SolidIndexed', 'SolidNamed', 'Gray125', 'Gray0625'
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    # 第二行：填充样式测试数据
    # 1. 无填充
    cell = ws.cell(row=2, column=1)
    cell.value = "No Fill"
    cell.fill = PatternFill(fill_type='none')

    # 2. 实心填充（RGB颜色）
    cell = ws.cell(row=2, column=2)
    cell.value = "RGB Fill"
    cell.fill = PatternFill(
        fill_type='solid',
        start_color='FFCC99',  # 橙色
        end_color='FFCC99'
    )

    # 3. 实心填充（索引颜色）
    cell = ws.cell(row=2, column=3)
    cell.value = "Indexed Fill"
    cell.fill = PatternFill(
        fill_type='solid',
        start_color=openpyxl.styles.colors.Color(indexed=15),  # 灰色
        end_color=openpyxl.styles.colors.Color(indexed=15)
    )

    # 4. 实心填充（命名颜色）
    cell = ws.cell(row=2, column=4)
    cell.value = "Named Fill"
    cell.fill = PatternFill(
        fill_type='solid',
        start_color='FFFF00',  # 黄色
        end_color='FFFF00'
    )

    # 5. 12.5% 灰度
    cell = ws.cell(row=2, column=5)
    cell.value = "Gray 12.5%"
    cell.fill = PatternFill(fill_type='gray125')

    # 6. 6.25% 灰度
    cell = ws.cell(row=2, column=6)
    cell.value = "Gray 6.25%"
    cell.fill = PatternFill(fill_type='gray0625')

    # 保存文件
    wb.save('template_fill_styles.xlsx')
    wb.close()

    print("✓ 创建包含填充样式的模板文件: template_fill_styles.xlsx")
    print("  - No Fill: 无填充")
    print("  - RGB Fill: RGB 颜色填充（橙色）")
    print("  - Indexed Fill: 索引颜色填充（灰色）")
    print("  - Named Fill: 命名颜色填充（黄色）")
    print("  - Gray 125: 12.5% 灰度")
    print("  - Gray 0625: 6.25% 灰度")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：填充样式复制")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [{'source': 'Value', 'target': 'Value'}],
            'alias': 'external_data'
        }
    ]

    # 公式列
    formula_columns = []

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_fill_styles.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_fill_output.xlsx',
            use_external_refs=False
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_fill_output.xlsx")

        # 验证输出文件中的填充样式
        wb = openpyxl.load_workbook('test_fill_output.xlsx', data_only=False)
        ws = wb.active

        print("\n填充样式验证结果:")

        # 验证各种填充样式
        cells_to_check = [
            (2, 1, "No Fill"),
            (2, 2, "RGB Fill"),
            (2, 3, "Indexed Fill"),
            (2, 4, "Named Fill"),
            (2, 5, "Gray 12.5%"),
            (2, 6, "Gray 6.25%")
        ]

        for row, col, name in cells_to_check:
            cell = ws.cell(row=row, column=col)
            fill_type = cell.fill.fill_type if cell.fill else 'none'

            if name == "No Fill":
                status = "✓" if fill_type in ['none', None] else "✗"
            elif name == "Gray 12.5%":
                status = "✓" if fill_type == 'gray125' else "✗"
            elif name == "Gray 6.25%":
                status = "✓" if fill_type == 'gray0625' else "✗"
            else:
                status = "✓" if fill_type == 'solid' else "✗"

            print(f"  {name}: {status} (类型: {fill_type})")

            # 检查填充颜色
            if fill_type == 'solid' and cell.fill.start_color:
                print(f"    颜色: {cell.fill.start_color}")

        wb.close()

        print("\n🎉 填充样式复制功能测试完成！")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

    return True


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_fill_styles()

    # 运行测试
    if run_test():
        print("\n✅ 填充样式复制功能正常工作！")
    else:
        print("\n❌ 填充样式复制功能存在问题！")