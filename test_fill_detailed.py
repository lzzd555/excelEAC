"""
详细测试各种填充类型的创建和复制
"""
import openpyxl
from openpyxl.styles import PatternFill
import os

# 创建测试文件
def test_fill_creation():
    """测试各种填充类型的创建方式"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TestSheet'

    # 测试1: 无填充
    print("\n=== 测试1: 无填充 ===")
    cell1 = ws.cell(row=1, column=1)
    cell1.value = "No Fill"
    cell1.fill = PatternFill(fill_type='none')
    print(f"设置后: fill_type={cell1.fill.fill_type if cell1.fill else 'None'}")

    # 测试2: RGB填充 - 方法A (字符串)
    print("\n=== 测试2: RGB填充 - 方法A ===")
    cell2 = ws.cell(row=2, column=1)
    cell2.value = "RGB Fill A"
    cell2.fill = PatternFill(fill_type='solid', start_color='FFCC99')
    print(f"设置后: fill_type={cell2.fill.fill_type}, start_color={cell2.fill.start_color}")

    # 测试3: RGB填充 - 方法B (Color对象)
    print("\n=== 测试3: RGB填充 - 方法B ===")
    from openpyxl.styles.colors import Color
    cell3 = ws.cell(row=3, column=1)
    cell3.value = "RGB Fill B"
    cell3.fill = PatternFill(fill_type='solid', start_color=Color(rgb='FFCC99'))
    print(f"设置后: fill_type={cell3.fill.fill_type}, start_color={cell3.fill.start_color}")

    # 测试4: 索引颜色填充
    print("\n=== 测试4: 索引颜色填充 ===")
    from openpyxl.styles.colors import Color
    cell4 = ws.cell(row=4, column=1)
    cell4.value = "Indexed Fill"
    cell4.fill = PatternFill(
        fill_type='solid',
        start_color=Color(indexed=15)  # 灰色
    )
    print(f"设置后: fill_type={cell4.fill.fill_type}, start_color={cell4.fill.start_color}")

    # 测试5: 灰度填充
    print("\n=== 测试5: 灰度填充 ===")
    cell5 = ws.cell(row=5, column=1)
    cell5.value = "Gray 125"
    cell5.fill = PatternFill(fill_type='gray125')
    print(f"设置后: fill_type={cell5.fill.fill_type}")

    # 测试6: 另一个灰度填充
    cell6 = ws.cell(row=6, column=1)
    cell6.value = "Gray 0625"
    cell6.fill = PatternFill(fill_type='gray0625')
    print(f"设置后: fill_type={cell6.fill.fill_type}")

    # 保存文件
    wb.save('test_fill_detailed.xlsx')
    wb.close()

    print("\n=== 重新加载并检查 ===")
    # 重新加载
    wb2 = openpyxl.load_workbook('test_fill_detailed.xlsx', data_only=False)
    ws2 = wb2.active

    for row in range(1, 7):
        cell = ws2.cell(row=row, column=1)
        print(f"\n行 {row}: '{cell.value}'")
        if cell.fill:
            print(f"  fill_type: {cell.fill.fill_type}")
            print(f"  start_color: {cell.fill.start_color}")
            print(f"  end_color: {cell.fill.end_color}")
        else:
            print("  fill: None")

    wb2.close()

def test_fill_copy_from_template():
    """测试从模板文件复制填充样式"""
    # 创建模板
    wb_template = openpyxl.Workbook()
    ws_template = wb_template.active
    ws_template.title = 'Template'

    # 设置各种填充
    ws_template.cell(row=1, column=1).value = "No Fill"
    ws_template.cell(row=1, column=1).fill = PatternFill(fill_type='none')

    ws_template.cell(row=1, column=2).value = "RGB Fill"
    ws_template.cell(row=1, column=2).fill = PatternFill(fill_type='solid', start_color='FFCC99')

    ws_template.cell(row=1, column=3).value = "Gray Fill"
    ws_template.cell(row=1, column=3).fill = PatternFill(fill_type='gray125')

    wb_template.save('template_fills.xlsx')
    wb_template.close()

    # 复制到新文件
    wb_copy = openpyxl.Workbook()
    ws_copy = wb_copy.active

    # 从模板读取并复制
    wb_load = openpyxl.load_workbook('template_fills.xlsx', data_only=False)
    ws_load = wb_load.active

    for col in range(1, 4):
        src_cell = ws_load.cell(row=1, column=col)
        dst_cell = ws_copy.cell(row=1, column=col)
        dst_cell.value = src_cell.value

        if src_cell.fill:
            print(f"\n复制列 {col}:")
            print(f"  原始: type={src_cell.fill.fill_type}")
            try:
                # 尝试复制
                dst_cell.fill = PatternFill(fill_type=src_cell.fill.fill_type)
                print(f"  ✓ 复制成功: type={dst_cell.fill.fill_type}")
            except Exception as e:
                print(f"  ✗ 复制失败: {e}")
                dst_cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF')
                print(f"  使用默认填充: type={dst_cell.fill.fill_type}")

    wb_copy.save('copy_fills.xlsx')
    wb_load.close()

if __name__ == "__main__":
    test_fill_creation()
    print("\n" + "="*50)
    test_fill_copy_from_template()

    # 清理
    for f in ['test_fill_detailed.xlsx', 'template_fills.xlsx', 'copy_fills.xlsx']:
        if os.path.exists(f):
            os.remove(f)