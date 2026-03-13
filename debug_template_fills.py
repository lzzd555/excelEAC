"""
调试模板文件中的填充类型
"""
import openpyxl

# 读取模板文件
wb = openpyxl.load_workbook('template_fill_styles.xlsx', data_only=False)
ws = wb.active

print("模板文件填充类型分析:")
print("=" * 50)

# 检查每个单元格的填充
headers = ['NoFill', 'SolidRGB', 'SolidIndexed', 'SolidNamed', 'Gray125', 'Gray0625']

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    print(f"\n列 {col_idx}: '{header}'")

    # 检查标题行
    header_cell = ws.cell(row=1, column=col_idx)
    print(f"  标题单元格 (行1):")
    if header_cell.fill:
        print(f"    fill_type: {header_cell.fill.fill_type}")
        print(f"    has_style: {header_cell.has_style}")
    else:
        print(f"    fill: None")

    # 检查数据行
    data_cell = ws.cell(row=2, column=col_idx)
    print(f"  数据单元格 (行2):")
    if data_cell.fill:
        print(f"    fill_type: {data_cell.fill.fill_type}")
        print(f"    start_color: {data_cell.fill.start_color}")
        if hasattr(data_cell.fill.start_color, 'rgb'):
            print(f"    rgb: {data_cell.fill.start_color.rgb}")
        if hasattr(data_cell.fill.start_color, 'indexed'):
            print(f"    indexed: {data_cell.fill.start_color.indexed}")
    else:
        print(f"    fill: None")

wb.close()