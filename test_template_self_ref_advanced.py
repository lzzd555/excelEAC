"""
高级测试：测试多种格式的模板自引用
"""
import pandas as pd
import openpyxl
import os

# 清理旧文件
for f in ['external_data.xlsx', 'template_with_sheet_ref.xlsx', 'test_sheet_ref_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Sales': [100, 200, 300, 400, 500],
        'Cost': [50, 100, 150, 200, 250],
        'Tax': [10, 20, 30, 40, 50]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含 sheet 名称引用的模板文件
def create_template_with_sheet_ref():
    """创建模板文件，使用带 sheet 名称的引用"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MyTemplateSheet'  # 设置模板 sheet 名称

    # 第一行：列名
    headers = ['Sales', 'Cost', 'Tax', 'SubTotal', 'Profit', 'Ratio']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # 第二行：公式模板（使用带 sheet 名称的引用）
    # SubTotal: =MyTemplateSheet!A2-MyTemplateSheet!B2-MyTemplateSheet!C2（带 sheet 名称的自引用）
    ws.cell(row=2, column=4).value = "=MyTemplateSheet!A2-MyTemplateSheet!B2-MyTemplateSheet!C2"

    # Profit: =MyTemplateSheet!D2*0.2（带 sheet 名称的自引用）
    ws.cell(row=2, column=5).value = "=MyTemplateSheet!D2*0.2"

    # Ratio: =MyTemplateSheet!E2/MyTemplateSheet!D2（带 sheet 名称的自引用）
    ws.cell(row=2, column=6).value = "=MyTemplateSheet!E2/MyTemplateSheet!D2"

    # 保存文件
    wb.save('template_with_sheet_ref.xlsx')
    wb.close()

    print("✓ 创建模板文件: template_with_sheet_ref.xlsx")
    print("  - Sheet 名称: MyTemplateSheet")
    print("  - 包含带 sheet 名称的自引用公式:")
    print("    - SubTotal = MyTemplateSheet!A2 - MyTemplateSheet!B2 - MyTemplateSheet!C2")
    print("    - Profit = MyTemplateSheet!D2 * 0.2")
    print("    - Ratio = MyTemplateSheet!E2 / MyTemplateSheet!D2")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：带 sheet 名称的模板自引用")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [
                {'source': 'Sales', 'target': 'Sales'},
                {'source': 'Cost', 'target': 'Cost'},
                {'source': 'Tax', 'target': 'Tax'}
            ],
            'alias': 'external_data'
        }
    ]

    # 公式列（这些列的公式会被应用到输出文件）
    formula_columns = ['SubTotal', 'Profit', 'Ratio']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_with_sheet_ref.xlsx',
            template_sheet='MyTemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_sheet_ref_output.xlsx',
            use_external_refs=True  # 使用外部引用模式
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_sheet_ref_output.xlsx")
        print("\n实际结果:")
        print("  SubTotal: =MyTemplateSheet!A2-MyTemplateSheet!B2-MyTemplateSheet!C2")

        # 验证输出文件中的公式
        print("\n正在验证输出文件中的公式...")
        wb = openpyxl.load_workbook('test_sheet_ref_output.xlsx', data_only=False)
        ws = wb.active

        # 读取第二行的公式
        subtotal_formula = ws.cell(row=2, column=4).value
        profit_formula = ws.cell(row=2, column=5).value
        ratio_formula = ws.cell(row=2, column=6).value

        print(f"  SubTotal: {subtotal_formula}")
        print(f"  Profit: {profit_formula}")
        print(f"  Ratio: {ratio_formula}")

        wb.close()

        # 检查是否包含输出文件名（不应该包含）
        if 'test_sheet_ref_output.xlsx' in str(subtotal_formula):
            print("\n✗ 错误：公式中包含输出文件名（应该是本地引用）")
            print(f"  发现: {subtotal_formula}")
        else:
            print("\n✓ 正确：公式使用本地引用格式")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_sheet_ref()

    # 运行测试
    run_test()
