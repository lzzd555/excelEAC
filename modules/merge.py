"""
Excel表合并模块
提供基于指定列匹配的数据合并功能
"""

import pandas as pd
from typing import List, Dict, Any, Optional, Union
import os


def merge_excel_tables(
    table_a_file: str,
    table_a_sheet: str,
    table_b_file: str,
    table_b_sheet: str,
    match_columns: Union[List[str], Dict[str, str]],
    table_a_extra_columns: Optional[List[str]] = None,
    table_b_extra_columns: Optional[List[str]] = None,
    output_file: str = 'merge_result.xlsx',
    string_columns: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    将两张Excel表中指定列数据相同的行合并成新表

    参数:
        table_a_file: 表A的Excel文件路径
        table_a_sheet: 表A的工作表名称
        table_b_file: 表B的Excel文件路径
        table_b_sheet: 表B的工作表名称
        match_columns: 需要匹配的列名。支持两种格式：
            - List[str]: ['ID', '部门'] - 表A和表B使用相同列名
            - Dict[str, str]: {'ID': '员工编号', '部门': '部门编码'} - 表A列名映射到表B列名
        table_a_extra_columns: 从表A中额外添加的列名列表（除匹配列外）
        table_b_extra_columns: 从表B中额外添加的列名列表（除匹配列外）
        output_file: 输出文件名
        string_columns: 需要保持为字符串格式的列名列表（避免"001"变成1）

    返回:
        合并后的DataFrame

    示例:
        旧格式（相同列名）:
            match_columns=['ID', '部门']
        新格式（不同列名映射）:
            match_columns={'ID': '员工编号', '部门': '部门编码'}
        从表A中额外加入c,d
        从表B中额外加入e,f,g
        最后新表中的列为匹配列+c+d+e+f+g
    """
    # 确保输出文件在当前文件夹
    if not os.path.isabs(output_file):
        # 如果是相对路径，确保输出到当前文件夹
        output_file = os.path.join(os.getcwd(), output_file)

    print("=== Excel表合并工具 ===\n")

    # 向后兼容处理：将 List[str] 转换为 Dict[str, str]
    if isinstance(match_columns, list):
        match_columns = {col: col for col in match_columns}

    # 1. 读取表A数据
    print(f"正在读取表A: {table_a_file} 的 {table_a_sheet} 工作表...")
    dtype_dict_a = {}
    if string_columns:
        for col in string_columns:
            dtype_dict_a[col] = 'string'

    df_a = pd.read_excel(table_a_file, sheet_name=table_a_sheet, dtype=dtype_dict_a)

    # 确保字符串列保持字符串格式
    if string_columns:
        for col in string_columns:
            if col in df_a.columns:
                df_a[col] = df_a[col].astype('string')

    print(f"表A数据: {len(df_a)} 行, {len(df_a.columns)} 列")
    print(f"表A列名: {list(df_a.columns)}\n")

    # 2. 读取表B数据
    print(f"正在读取表B: {table_b_file} 的 {table_b_sheet} 工作表...")
    dtype_dict_b = {}
    if string_columns:
        # 注意：表B的列名可能与表A不同，需要根据match_columns映射
        for a_col, b_col in match_columns.items():
            if a_col in string_columns and b_col not in dtype_dict_b:
                dtype_dict_b[b_col] = 'string'

    df_b = pd.read_excel(table_b_file, sheet_name=table_b_sheet, dtype=dtype_dict_b)

    # 确保字符串列保持字符串格式
    if string_columns:
        for a_col, b_col in match_columns.items():
            if a_col in string_columns and b_col in df_b.columns:
                df_b[b_col] = df_b[b_col].astype('string')

    print(f"表B数据: {len(df_b)} 行, {len(df_b.columns)} 列")
    print(f"表B列名: {list(df_b.columns)}\n")

    # 3. 验证匹配列是否存在
    print("正在验证匹配列...")

    # 检查表A是否有所有key
    missing_in_a = [col for col in match_columns.keys() if col not in df_a.columns]
    if missing_in_a:
        raise ValueError(f"表A中不存在匹配列: {missing_in_a}")

    # 检查表B是否有所有对应的value
    missing_in_b = [b_col for a_col, b_col in match_columns.items() if b_col not in df_b.columns]
    if missing_in_b:
        raise ValueError(f"表B中不存在匹配列: {missing_in_b}")

    print(f"列映射关系: {match_columns}")
    print(f"表A额外列: {table_a_extra_columns if table_a_extra_columns else '无'}")
    print(f"表B额外列: {table_b_extra_columns if table_b_extra_columns else '无'}\n")

    # 4. 验证额外列是否存在
    if table_a_extra_columns:
        missing_extra_a = [col for col in table_a_extra_columns if col not in df_a.columns]
        if missing_extra_a:
            raise ValueError(f"表A额外列不存在: {missing_extra_a}")

    if table_b_extra_columns:
        missing_extra_b = [col for col in table_b_extra_columns if col not in df_b.columns]
        if missing_extra_b:
            raise ValueError(f"表B额外列不存在: {missing_extra_b}")

    # 5. 确定最终列顺序
    # 匹配列（使用表A的列名）+ 表A额外列 + 表B额外列
    final_columns = list(match_columns.keys())

    # 添加表A的额外列（去除重复）
    if table_a_extra_columns:
        for col in table_a_extra_columns:
            if col not in final_columns:
                final_columns.append(col)

    # 添加表B的额外列（去除重复）
    if table_b_extra_columns:
        for col in table_b_extra_columns:
            if col not in final_columns:
                final_columns.append(col)

    print(f"最终输出列: {final_columns}")
    print(f"共 {len(final_columns)} 列\n")

    # 6. 执行表合并
    print("正在执行表合并...")

    # 如果没有匹配列，返回空DataFrame
    if not match_columns:
        print("没有匹配列，返回空DataFrame\n")
        merged_df = pd.DataFrame(columns=final_columns)
        print(f"合并结果: {len(merged_df)} 行, {len(merged_df.columns)} 列")

        # 输出到Excel
        print(f"正在输出结果到 {output_file}...")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            merged_df.to_excel(writer, sheet_name='合并结果', index=False)

        print(f"文件已保存到: {output_file}")
        return merged_df

    merged_rows = []
    merged_count = 0

    # 遍历表A的每一行
    for _, row_a in df_a.iterrows():
        # 在表B中查找匹配的行
        match_condition = True
        for a_col, b_col in match_columns.items():
            match_condition = match_condition & (df_b[b_col] == row_a[a_col])

        # 找到所有匹配的行
        matched_rows_b = df_b[match_condition]

        if not matched_rows_b.empty:
            # 对每个匹配的表B行，创建合并行
            for _, row_b in matched_rows_b.iterrows():
                # 创建新行字典
                new_row = {}

                # 添加匹配列的值（使用表A的列名）
                for a_col in match_columns.keys():
                    new_row[a_col] = row_a[a_col]

                # 添加表A的额外列
                if table_a_extra_columns:
                    for col in table_a_extra_columns:
                        if col in df_a.columns:
                            new_row[col] = row_a[col]

                # 添加表B的额外列
                if table_b_extra_columns:
                    for col in table_b_extra_columns:
                        if col in df_b.columns:
                            new_row[col] = row_b[col]

                merged_rows.append(new_row)
                merged_count += 1

    print(f"合并完成: 共生成 {merged_count} 行数据\n")

    # 7. 创建合并后的DataFrame
    # 确保列顺序正确
    if merged_rows:
        merged_df = pd.DataFrame(merged_rows)

        # 按最终列顺序重新排序列
        # 只保留最终列中存在的列
        available_columns = [col for col in final_columns if col in merged_df.columns]
        merged_df = merged_df[available_columns]
    else:
        # 如果没有匹配的行，仍然创建具有正确列结构的空DataFrame
        merged_df = pd.DataFrame(columns=final_columns)

    print(f"合并结果: {len(merged_df)} 行, {len(merged_df.columns)} 列")

    # 8. 输出到Excel
    print(f"正在输出结果到 {output_file}...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name='合并结果', index=False)

        # 为字符串列设置格式
        if string_columns:
            from openpyxl.utils import get_column_letter
            ws = writer.sheets['合并结果']

            for col_idx, col_name in enumerate(merged_df.columns):
                if col_name in string_columns:
                    # 设置列宽
                    ws.column_dimensions[get_column_letter(col_idx + 1)].width = 15

                    # 设置整个列的格式为文本
                    for row in range(1, len(merged_df) + 2):  # 包含标题行
                        cell = ws.cell(row=row, column=col_idx + 1)
                        # 设置单元格格式为文本，保持前导零
                        cell.number_format = '@'

                    # 确保数据是字符串格式
                    for row_idx in range(len(merged_df)):
                        row_idx_excel = row_idx + 2
                        cell_value = merged_df.iloc[row_idx, col_idx]
                        if pd.notna(cell_value):
                            ws.cell(row=row_idx_excel, column=col_idx + 1).value = str(cell_value)

    print(f"文件已保存到: {output_file}")
    return merged_df


# 主函数入口
if __name__ == "__main__":
    # 示例使用
    print("=== Excel表合并工具示例 ===\n")

    # 创建示例数据
    table_a_data = {
        'ID': ['A001', 'A002', 'A003', 'A004'],
        '姓名': ['张三', '李四', '王五', '赵六'],
        '部门': ['销售部', '销售部', '市场部', '技术部'],
        '年龄': [25, 30, 28, 35],
        '入职日期': ['2020-01-01', '2020-02-01', '2020-03-01', '2020-04-01']
    }
    df_a = pd.DataFrame(table_a_data)
    df_a.to_excel('example_table_a.xlsx', index=False)

    table_b_data = {
        'ID': ['A001', 'A002', 'A003', 'B001'],
        '职位': ['经理', '专员', '主管', '经理'],
        '薪资': [10000, 8000, 12000, 15000],
        '绩效等级': ['A', 'B', 'A', 'A']
    }
    df_b = pd.DataFrame(table_b_data)
    df_b.to_excel('example_table_b.xlsx', index=False)

    print("示例数据已创建: example_table_a.xlsx, example_table_b.xlsx\n")

    # 执行合并（使用相同的列名）
    result1 = merge_excel_tables(
        table_a_file='example_table_a.xlsx',
        table_a_sheet='Sheet1',
        table_b_file='example_table_b.xlsx',
        table_b_sheet='Sheet1',
        match_columns=['ID'],  # 按ID列匹配（旧格式，仍然支持）
        table_a_extra_columns=['姓名', '部门', '年龄', '入职日期'],  # 表A额外列
        table_b_extra_columns=['职位', '薪资', '绩效等级'],  # 表B额外列
        output_file='example_merge_result1.xlsx',
        string_columns=['ID']  # ID列保持字符串格式
    )

    print("\n=== 使用不同列名映射的示例 ===\n")

    # 创建不同列名的示例数据
    table_b_data_diff = {
        '员工编号': ['A001', 'A002', 'A003', 'B001'],
        '岗位': ['经理', '专员', '主管', '经理'],
        '薪酬': [10000, 8000, 12000, 15000],
        '评级': ['A', 'B', 'A', 'A']
    }
    df_b_diff = pd.DataFrame(table_b_data_diff)
    df_b_diff.to_excel('example_table_b_diff.xlsx', index=False)

    # 执行合并（使用不同的列名映射）
    result2 = merge_excel_tables(
        table_a_file='example_table_a.xlsx',
        table_a_sheet='Sheet1',
        table_b_file='example_table_b_diff.xlsx',
        table_b_sheet='Sheet1',
        match_columns={'ID': '员工编号', '部门': '岗位'},  # 列名映射（新格式）
        table_a_extra_columns=['姓名', '年龄', '入职日期'],  # 表A额外列
        table_b_extra_columns=['薪酬', '评级'],  # 表B额外列
        output_file='example_merge_result2.xlsx',
        string_columns=['ID', '员工编号']  # 保持字符串格式
    )

    print("\n合并示例完成！")

    print("\n合并结果:")
    print(result)
    print("\n合并完成！")
