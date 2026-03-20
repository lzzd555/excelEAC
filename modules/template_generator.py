"""
Excel模板生成器模块
基于模板生成Excel，支持多数据源、公式保留和列映射
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
import os
import re
import copy


# ==================== 数据类定义 ====================

@dataclass
class DataColumnMapping:
    """列映射配置"""
    source_column: str  # 源列名
    target_column: str  # 目标列名（与模板列名匹配）


@dataclass
class DataSourceConfig:
    """数据源配置"""
    file_path: str                           # 数据文件路径
    sheet_name: str                          # sheet名称
    column_mappings: List[DataColumnMapping] # 列映射集合
    alias: str = ""                          # 别名（如"sheet0", "sheet1"）


@dataclass
class CellStyle:
    """单元格样式"""
    font: Optional[Dict[str, Any]] = None
    fill: Optional[Dict[str, Any]] = None
    border: Optional[Dict[str, Any]] = None
    alignment: Optional[Dict[str, Any]] = None
    number_format: Optional[str] = None
    protection: Optional[Dict[str, Any]] = None


# ==================== 颜色处理函数 ====================

def copy_color(source_color):
    """
    复制颜色对象，支持 RGB、主题、索引和自动颜色

    Args:
        source_color: 源颜色对象 (openpyxl.styles.colors.Color)

    Returns:
        新的 Color 对象，或原始颜色值（RGB字符串）
    """
    from openpyxl.styles.colors import Color

    if source_color is None:
        return None

    color_type = getattr(source_color, 'type', None)

    if color_type == 'rgb':
        return _copy_rgb_color(source_color)
    elif color_type == 'theme':
        return _copy_theme_color(source_color)
    elif color_type == 'indexed':
        return _copy_indexed_color(source_color)
    elif color_type == 'auto' or (hasattr(source_color, 'auto') and source_color.auto):
        return Color(auto=True)

    return source_color


def _copy_rgb_color(source_color):
    """复制 RGB 颜色"""
    from openpyxl.styles.colors import Color
    if hasattr(source_color, 'rgb') and isinstance(source_color.rgb, str) and source_color.rgb:
        return source_color.rgb
    return source_color


def _copy_theme_color(source_color):
    """复制主题颜色"""
    from openpyxl.styles.colors import Color
    try:
        theme_val = int(source_color.theme)
        tint_val = float(source_color.tint) if source_color.tint else 0
        return Color(theme=theme_val, tint=tint_val)
    except (TypeError, ValueError):
        return source_color


def _copy_indexed_color(source_color):
    """复制索引颜色"""
    from openpyxl.styles.colors import Color
    try:
        indexed_val = int(source_color.indexed)
        return Color(indexed=indexed_val)
    except (TypeError, ValueError):
        return source_color


def copy_side(source_side):
    """
    复制边框线样式（Side对象），正确处理颜色

    Args:
        source_side: 源边框线对象

    Returns:
        新的 Side 对象
    """
    if source_side is None or source_side.border_style is None:
        return None

    color = copy_color(source_side.color) if source_side.color else None

    return Side(
        style=source_side.border_style,
        color=color
    )


# ==================== 单元格样式复制函数 ====================

def copy_cell_style(source_cell, target_cell) -> None:
    """
    复制单元格样式

    Args:
        source_cell: 源单元格
        target_cell: 目标单元格
    """
    if not source_cell.has_style:
        return

    _copy_font_style(source_cell, target_cell)
    _copy_fill_style(source_cell, target_cell)
    _copy_border_style(source_cell, target_cell)
    _copy_alignment_style(source_cell, target_cell)

    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format

    if source_cell.protection:
        target_cell.protection = Protection(
            locked=source_cell.protection.locked,
            hidden=source_cell.protection.hidden
        )


def _copy_font_style(source_cell, target_cell) -> None:
    """复制字体样式"""
    if not source_cell.font:
        return

    font_args = {
        'name': source_cell.font.name,
        'size': source_cell.font.size,
        'bold': source_cell.font.bold,
        'italic': source_cell.font.italic,
        'vertAlign': source_cell.font.vertAlign,
        'underline': source_cell.font.underline,
        'strike': source_cell.font.strike,
    }

    if source_cell.font.color:
        font_args['color'] = _get_font_color(source_cell.font.color)

    target_cell.font = Font(**font_args)


def _get_font_color(src_color):
    """获取字体颜色"""
    from openpyxl.styles.colors import Color
    color_type = getattr(src_color, 'type', None)

    if color_type == 'rgb' and hasattr(src_color, 'rgb'):
        if isinstance(src_color.rgb, str) and src_color.rgb:
            return Color(rgb=src_color.rgb)
    elif color_type == 'theme' and hasattr(src_color, 'theme'):
        try:
            return Color(theme=int(src_color.theme), tint=float(src_color.tint) if src_color.tint else 0)
        except (TypeError, ValueError):
            pass
    elif color_type == 'indexed' and hasattr(src_color, 'indexed'):
        try:
            return Color(indexed=int(src_color.indexed))
        except (TypeError, ValueError):
            pass
    elif color_type == 'auto' or (hasattr(src_color, 'auto') and src_color.auto):
        return Color(auto=True)

    return src_color


def _copy_fill_style(source_cell, target_cell) -> None:
    """复制填充样式"""
    if not source_cell.fill:
        return

    try:
        fill_type = source_cell.fill.fill_type
        _apply_fill_by_type(source_cell, target_cell, fill_type)
    except Exception as e:
        _apply_fallback_fill(target_cell, fill_type, e)


def _apply_fill_by_type(source_cell, target_cell, fill_type) -> None:
    """根据填充类型应用填充样式"""
    if fill_type is None or fill_type == 'none':
        target_cell.fill = PatternFill(fill_type='none')
    elif fill_type in ('gray125', 'gray0625'):
        target_cell.fill = PatternFill(fill_type=fill_type)
    elif fill_type == 'solid':
        _apply_solid_fill(source_cell, target_cell)
    else:
        target_cell.fill = PatternFill(fill_type=fill_type)


def _apply_solid_fill(source_cell, target_cell) -> None:
    """应用实心填充"""
    from openpyxl.styles.colors import Color
    start_color = source_cell.fill.start_color
    end_color = source_cell.fill.end_color

    if not start_color:
        target_cell.fill = PatternFill(fill_type='none')
        return

    color_value = _get_fill_color_value(start_color, Color)

    if color_value is not None:
        target_cell.fill = PatternFill(
            fill_type='solid',
            start_color=color_value,
            end_color=color_value
        )
    else:
        target_cell.fill = PatternFill(
            fill_type='solid',
            start_color=start_color,
            end_color=end_color if end_color else start_color
        )


def _get_fill_color_value(start_color, Color):
    """获取填充颜色值"""
    if hasattr(start_color, 'rgb') and isinstance(start_color.rgb, str) and start_color.rgb:
        return start_color.rgb
    elif hasattr(start_color, 'theme') and start_color.theme is not None:
        return Color(theme=start_color.theme, tint=start_color.tint or 0)
    elif hasattr(start_color, 'indexed') and start_color.indexed is not None:
        return Color(indexed=start_color.indexed)
    return None


def _apply_fallback_fill(target_cell, fill_type, error) -> None:
    """应用后备填充样式"""
    try:
        if fill_type is None:
            target_cell.fill = PatternFill(fill_type='none')
        else:
            target_cell.fill = PatternFill(fill_type=fill_type)
    except Exception:
        print(f"⚠️ 跳过填充样式复制（错误: {error}）")


def _copy_border_style(source_cell, target_cell) -> None:
    """复制边框样式"""
    if not source_cell.border:
        return

    target_cell.border = Border(
        left=copy_side(source_cell.border.left),
        right=copy_side(source_cell.border.right),
        top=copy_side(source_cell.border.top),
        bottom=copy_side(source_cell.border.bottom),
        diagonal=copy_side(source_cell.border.diagonal),
        diagonal_direction=source_cell.border.diagonal_direction,
        outline=source_cell.border.outline,
        horizontal=copy_side(source_cell.border.horizontal),
        vertical=copy_side(source_cell.border.vertical)
    )


def _copy_alignment_style(source_cell, target_cell) -> None:
    """复制对齐样式"""
    if not source_cell.alignment:
        return

    target_cell.alignment = Alignment(
        horizontal=source_cell.alignment.horizontal,
        vertical=source_cell.alignment.vertical,
        text_rotation=source_cell.alignment.text_rotation,
        wrap_text=source_cell.alignment.wrap_text,
        shrink_to_fit=source_cell.alignment.shrink_to_fit,
        indent=source_cell.alignment.indent
    )


# ==================== 外部链接读取函数 ====================

def read_external_links(xlsx_file: str) -> Dict[int, str]:
    """
    读取Excel文件中的外部链接映射

    Args:
        xlsx_file: Excel文件路径

    Returns:
        Dict[int, str]: 外部链接映射 {索引号: 文件名}
    """
    links = {}

    try:
        with openpyxl.load_workbook(xlsx_file, data_only=False) as wb:
            links = _extract_external_links(wb)
    except Exception as e:
        print(f"   警告: 无法读取外部链接: {e}")

    return links


def _extract_external_links(wb) -> Dict[int, str]:
    """从工作簿中提取外部链接"""
    links = {}

    if not hasattr(wb, 'external_links') or not wb.external_links:
        return links

    for link in wb.external_links:
        link_info = _parse_single_link(link)
        if link_info:
            links.update(link_info)

    return links


def _parse_single_link(link) -> Optional[Dict[int, str]]:
    """解析单个外部链接"""
    try:
        link_id = getattr(link, 'id', None)
        target = getattr(link, 'target', None) or getattr(link, 'file_link', None)

        if link_id is not None and target:
            if isinstance(link_id, int):
                return {link_id: target}

        # 尝试从字符串表示中提取信息
        link_str = str(link)
        id_match = re.search(r'id=(\d+)', link_str)
        target_match = re.search(r"target='([^']+)'", link_str)

        if id_match and target_match:
            return {int(id_match.group(1)): target_match.group(1)}

    except Exception:
        pass

    return None


# ==================== 模板结构读取函数 ====================

def read_template_structure(
    template_file: str,
    template_sheet: str
) -> Tuple[List[str], Dict[str, str], openpyxl.worksheet.worksheet.Worksheet, openpyxl.Workbook]:
    """
    读取模板的结构信息

    Args:
        template_file: 模板文件路径
        template_sheet: 模板sheet名称

    Returns:
        Tuple: (列名列表, 公式模板字典, 模板工作表对象, 工作簿对象)
               注意：调用者需要负责关闭返回的工作簿对象
    """
    wb = openpyxl.load_workbook(template_file, data_only=False)

    try:
        if template_sheet not in wb.sheetnames:
            raise ValueError(f"模板中不存在工作表: {template_sheet}")

        ws = wb[template_sheet]

        # 读取第一行作为列名
        columns = _read_template_columns(ws)

        # 读取第二行的公式（作为公式模板）
        formula_templates = _read_formula_templates(ws, columns)

        print(f"正在读取模板: {os.path.basename(template_file)} 的 {template_sheet} 工作表...")
        print(f"模板列名: {columns}")
        print(f"公式列: {[k for k, v in formula_templates.items() if v]}")

        return columns, formula_templates, ws, wb
    except Exception:
        wb.close()
        raise


def _read_template_columns(ws) -> List[str]:
    """读取模板列名"""
    columns = []
    for cell in ws[1]:
        if cell.value:
            columns.append(str(cell.value))
    return columns


def _read_formula_templates(ws, columns: List[str]) -> Dict[str, str]:
    """读取公式模板"""
    formula_templates = {}

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula_templates[col_name] = cell.value

    return formula_templates


# ==================== 公式替换辅助函数 ====================

def replace_link_indices_with_filenames(formula: str, link_mapping: Dict[int, str]) -> str:
    """
    将公式中的链接索引号替换为实际文件名

    Args:
        formula: 原始公式字符串
        link_mapping: 链接映射 {索引号: 文件名}

    Returns:
        str: 替换后的公式
    """
    if not link_mapping:
        return formula

    # 匹配 [数字] 格式的链接索引
    pattern = r'\[(\d+)\]'

    def replace_index(match):
        index = int(match.group(1))
        if index in link_mapping:
            filename = os.path.basename(link_mapping[index])
            return f'[{filename}]'
        return match.group(0)

    return re.sub(pattern, replace_index, formula)


# ==================== 数据源读取函数 ====================

def read_data_source(config: DataSourceConfig, string_columns: Optional[List[str]] = None) -> pd.DataFrame:
    """
    读取数据源文件

    Args:
        config: 数据源配置
        string_columns: 字符串列列表

    Returns:
        pd.DataFrame: 读取的数据
    """
    print(f"正在读取数据源: {os.path.basename(config.file_path)} 的 {config.sheet_name} 工作表...")

    # 读取所有列为字符串，避免自动类型转换
    df = pd.read_excel(
        config.file_path,
        sheet_name=config.sheet_name,
        dtype=str
    )

    # 应用列映射
    df = _apply_column_mappings(df, config.column_mappings)

    # 处理字符串列
    if string_columns:
        df = _process_string_columns(df, string_columns)

    print(f"数据源 '{config.alias}' 数据: {len(df)} 行, {len(df.columns)} 列")

    return df


def _apply_column_mappings(df: pd.DataFrame, mappings: List[DataColumnMapping]) -> pd.DataFrame:
    """应用列映射"""
    column_map = {m.source_column: m.target_column for m in mappings}
    renamed_columns = {}

    for col in df.columns:
        if col in column_map:
            renamed_columns[col] = column_map[col]

    return df.rename(columns=renamed_columns)


def _process_string_columns(df: pd.DataFrame, string_columns: List[str]) -> pd.DataFrame:
    """处理字符串列"""
    for col in string_columns:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: str(x) if pd.notna(x) else x)
    return df


# ==================== 数据合并函数 ====================

def merge_data_by_row(
    data_sources: List[Tuple[str, pd.DataFrame]],
    template_columns: List[str]
) -> pd.DataFrame:
    """
    按行号对齐合并多个数据源

    Args:
        data_sources: 数据源列表 [(别名, DataFrame), ...]
        template_columns: 模板列名列表

    Returns:
        pd.DataFrame: 合并后的数据
    """
    if not data_sources:
        return pd.DataFrame(columns=template_columns)

    # 找出最大行数
    max_rows = max(len(df) for _, df in data_sources)
    print(f"合并数据: 最大行数 = {max_rows}")

    # 创建合并后的DataFrame
    merged_data = {col: [None] * max_rows for col in template_columns}

    for alias, df in data_sources:
        for col in template_columns:
            if col in df.columns:
                for i, value in enumerate(df[col]):
                    if i < max_rows:
                        merged_data[col][i] = value

    result = pd.DataFrame(merged_data)
    print(f"合并完成: {len(result)} 行, {len(result.columns)} 列")

    return result


# ==================== 公式引用解析函数 ====================

def parse_formula_references(formula: str) -> List[Tuple[str, str, int]]:
    """
    解析公式中的引用

    Args:
        formula: 公式字符串

    Returns:
        List[Tuple]: [(sheet名, 列引用, 行号), ...]
    """
    references = []

    # 匹配单元格引用的正则模式
    patterns = [
        r"'([^']+)'!([A-Z]+)(\d+)",  # 'SheetName'!A1
        r"([A-Za-z_][A-Za-z0-9_]*)!([A-Z]+)(\d+)",  # SheetName!A1
    ]

    for pattern in patterns:
        matches = re.findall(pattern, formula)
        for match in matches:
            # 将行号转换为整数
            sheet, col, row = match
            references.append((sheet, col, int(row)))

    return references


# ==================== Sheet引用替换函数 ====================

# 正则模式常量
# 注意: 添加 \$? 来支持绝对引用（如 $A$1, $D:$D 等）
_QUOTED_PATTERN = r"'([^']+)'!(\$?[A-Z]+\$?\d*(?::\$?[A-Z]*\$?\d*)?)"
_BRACKET_PATTERN = r"(\[[^\]]+\][^!'\s]+)!(\$?[A-Z]+\$?\d*(?::\$?[A-Z]*\$?\d*)?)"
_UNQUOTED_PATTERN = r"([A-Za-z_][A-Za-z0-9_]*)!(\$?[A-Z]+\$?\d*(?::\$?[A-Z]*\$?\d*)?)"
_LOCAL_PATTERN = r"(?<![A-Za-z!'\"\\])(\$?[A-Z]+)(\$?\d+)(?![A-Za-z])"


def replace_sheet_references(
    formula: str,
    alias_to_info: Dict[str, Dict[str, str]],
    row_offset: int = 0,
    output_file_path: Optional[str] = None,
    external_links: Optional[Dict[int, str]] = None
) -> str:
    """
    替换公式中的sheet引用为外部文件引用，并调整行号
    """
    # 先处理带单引号的sheet名
    result = re.sub(_QUOTED_PATTERN,
                    lambda m: _replace_quoted_match(m, alias_to_info, row_offset, output_file_path, external_links),
                    formula, flags=re.IGNORECASE)

    # 处理带方括号但无单引号的格式
    result = re.sub(_BRACKET_PATTERN,
                    lambda m: _replace_bracket_match(m, alias_to_info, row_offset, output_file_path, external_links),
                    result, flags=re.IGNORECASE)

    # 最后处理不带单引号和方括号的sheet名
    result = re.sub(_UNQUOTED_PATTERN,
                    lambda m: _replace_unquoted_match(m, alias_to_info, row_offset, output_file_path),
                    result, flags=re.IGNORECASE)

    # 调整本地单元格引用的行号
    result = re.sub(_LOCAL_PATTERN,
                    lambda m: f"{m.group(1)}{int(m.group(2)) + row_offset}",
                    result)

    return result


def _adjust_cell_ref(cell_ref: str, row_offset: int) -> str:
    """调整单元格引用的行号"""
    if ':' in cell_ref:
        return _adjust_range_ref(cell_ref, row_offset)
    return _adjust_single_ref(cell_ref, row_offset)


def _adjust_single_ref(cell_ref: str, row_offset: int) -> str:
    """调整单个单元格引用"""
    col_match = re.match(r'([A-Z]+)(\d+)', cell_ref)
    if col_match:
        col = col_match.group(1)
        row = int(col_match.group(2)) + row_offset
        return f"{col}{row}"
    return cell_ref


def _adjust_range_ref(cell_ref: str, row_offset: int) -> str:
    """调整范围引用"""
    parts = cell_ref.split(':')
    adjusted_parts = []
    for part in parts:
        col_match = re.match(r'([A-Z]+)(\d*)', part)
        if col_match:
            col = col_match.group(1)
            row_str = col_match.group(2)
            if row_str:
                row = int(row_str) + row_offset
                adjusted_parts.append(f"{col}{row}")
            else:
                adjusted_parts.append(col)
        else:
            adjusted_parts.append(part)
    return ':'.join(adjusted_parts)


def _extract_sheet_name(full_reference: str) -> str:
    """从完整引用中提取实际的sheet名"""
    bracket_match = re.match(r'\[.+\](.+)', full_reference)
    if bracket_match:
        return bracket_match.group(1)
    return full_reference


def _find_matching_info(sheet_name: str, alias_to_info: Dict, external_links: Optional[Dict]) -> Optional[Dict]:
    """查找匹配的sheet信息"""
    actual_sheet_name = _extract_sheet_name(sheet_name)
    actual_sheet_name_lower = actual_sheet_name.lower()

    # 1. 精确匹配
    for key, info in alias_to_info.items():
        if key.lower() == actual_sheet_name_lower:
            return info

    # 2. 匹配实际的sheet名
    matching_infos = [
        (key, info) for key, info in alias_to_info.items()
        if info.get('sheet_name', '').lower() == actual_sheet_name_lower
    ]

    if len(matching_infos) > 1:
        return _resolve_multiple_matches(sheet_name, matching_infos)
    elif matching_infos:
        return matching_infos[0][1]

    # 3. 处理数字索引
    return _find_by_index(sheet_name, alias_to_info, external_links)


def _resolve_multiple_matches(sheet_name: str, matching_infos: List) -> Optional[Dict]:
    """解决多个匹配的情况"""
    filename_match = re.search(r'\[([^\]]+)\]', sheet_name)
    if filename_match:
        filename = filename_match.group(1).lower()
        for key, info in matching_infos:
            if filename in os.path.basename(info['file_path']).lower():
                return info
    return matching_infos[0][1]


def _find_by_index(sheet_name: str, alias_to_info: Dict, external_links: Optional[Dict]) -> Optional[Dict]:
    """通过索引查找"""
    bracket_match = re.match(r'\[(\d+)\]', sheet_name)
    if not bracket_match:
        return None

    index = bracket_match.group(1)

    if external_links and index in external_links:
        external_filename = external_links[index].lower()
        for key, info in alias_to_info.items():
            if external_filename in os.path.basename(info['file_path']).lower():
                return info

    return alias_to_info.get(index)


def _build_reference(info: Dict, adjusted_ref: str, output_file_path: Optional[str]) -> str:
    """构建引用字符串"""
    file_path = info.get('file_path', '')
    actual_sheet_name = info['sheet_name']

    # 检查是否为内部引用（数据源sheet已复制到输出文件）
    is_internal = info.get('is_internal', False)

    # 检查是否为本地引用
    is_local = (
        is_internal or
        (output_file_path and file_path and os.path.normpath(file_path) == os.path.normpath(output_file_path)) or
        info.get('is_template_self_reference', False)
    )

    if is_local:
        if any(c in actual_sheet_name for c in " -()&^%$#@!~`'\"\\"):
            return f"'{actual_sheet_name}'!{adjusted_ref}"
        return f"{actual_sheet_name}!{adjusted_ref}"

    file_name = os.path.basename(file_path)
    return f"'[{file_name}]{actual_sheet_name}'!{adjusted_ref}"


def _replace_quoted_match(match, alias_to_info: Dict, row_offset: int,
                          output_file_path: Optional[str], external_links: Optional[Dict]) -> str:
    """替换带引号的匹配"""
    full_reference = match.group(1)
    cell_ref = match.group(2).upper()

    info = _find_matching_info(full_reference, alias_to_info, external_links)
    if info:
        adjusted_ref = _adjust_cell_ref(cell_ref, row_offset)
        return _build_reference(info, adjusted_ref, output_file_path)
    return match.group(0)


def _replace_unquoted_match(match, alias_to_info: Dict, row_offset: int,
                            output_file_path: Optional[str]) -> str:
    """替换不带引号的匹配"""
    sheet_name = match.group(1)
    cell_ref = match.group(2).upper()

    info = _find_matching_info(sheet_name, alias_to_info, None)
    if info:
        adjusted_ref = _adjust_cell_ref(cell_ref, row_offset)
        return _build_reference(info, adjusted_ref, output_file_path)
    return match.group(0)


def _replace_bracket_match(match, alias_to_info: Dict, row_offset: int,
                           output_file_path: Optional[str], external_links: Optional[Dict]) -> str:
    """替换带方括号的匹配"""
    full_reference = match.group(1)
    cell_ref = match.group(2).upper()

    info = _find_matching_info(full_reference, alias_to_info, external_links)
    if info:
        adjusted_ref = _adjust_cell_ref(cell_ref, row_offset)
        return _build_reference(info, adjusted_ref, output_file_path)
    return match.group(0)


# ==================== 模板样式应用函数 ====================

def apply_template_styles(
    output_ws: openpyxl.worksheet.worksheet.Worksheet,
    template_ws: openpyxl.worksheet.worksheet.Worksheet,
    column_names: List[str],
    max_data_row: int
) -> None:
    """
    将模板的样式应用到输出文件

    Args:
        output_ws: 输出工作表对象
        template_ws: 模板工作表对象
        column_names: 列名列表
        max_data_row: 数据的最大行数
    """
    print("正在应用模板样式...")

    _apply_header_styles(output_ws, template_ws, column_names)
    _apply_data_styles(output_ws, template_ws, column_names, max_data_row)
    _copy_column_widths(output_ws, template_ws)
    _copy_row_height(output_ws, template_ws)

    print("✓ 模板样式应用完成")


def _apply_header_styles(output_ws, template_ws, column_names: List[str]) -> None:
    """应用标题行样式"""
    for col_idx, col_name in enumerate(column_names, start=1):
        template_cell = template_ws.cell(row=1, column=col_idx)
        output_cell = output_ws.cell(row=1, column=col_idx)

        if template_cell.has_style:
            copy_cell_style(template_cell, output_cell)


def _apply_data_styles(output_ws, template_ws, column_names: List[str], max_data_row: int) -> None:
    """应用数据行样式"""
    for col_idx, col_name in enumerate(column_names, start=1):
        template_cell = template_ws.cell(row=2, column=col_idx)

        for row_idx in range(2, max_data_row + 1):
            output_cell = output_ws.cell(row=row_idx, column=col_idx)

            if template_cell.has_style:
                copy_cell_style(template_cell, output_cell)


def _copy_column_widths(output_ws, template_ws) -> None:
    """复制列宽"""
    for col_idx in range(1, template_ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if template_ws.column_dimensions[col_letter].width:
            output_ws.column_dimensions[col_letter].width = template_ws.column_dimensions[col_letter].width


def _copy_row_height(output_ws, template_ws) -> None:
    """复制行高（标题行）"""
    if template_ws.row_dimensions[1].height:
        output_ws.row_dimensions[1].height = template_ws.row_dimensions[1].height


# ==================== 公式应用函数 ====================

def apply_formulas_to_output(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    formula_columns: List[str],
    formula_templates: Dict[str, str],
    alias_to_info: Dict[str, Dict[str, str]],
    start_row: int = 2,
    output_file_path: Optional[str] = None,
    external_links: Optional[Dict[int, str]] = None
) -> None:
    """
    将公式应用到输出文件

    Args:
        ws: 输出工作表对象
        formula_columns: 公式列名列表
        formula_templates: 公式模板字典
        alias_to_info: sheet名到文件信息的映射
        start_row: 开始行号
        output_file_path: 输出文件路径
        external_links: 外部链接映射
    """
    # 获取列名到列号的映射
    col_names = _get_column_names(ws)

    for col_name in formula_columns:
        if col_name not in formula_templates:
            continue

        formula = formula_templates[col_name]
        col_idx = col_names.get(col_name)

        if col_idx:
            _apply_formula_to_column(ws, col_idx, formula, alias_to_info,
                                    start_row, output_file_path, external_links, col_name)


def _get_column_names(ws) -> Dict[str, int]:
    """获取列名到列号的映射"""
    col_names = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            col_names[str(cell.value)] = col_idx
    return col_names


def _apply_formula_to_column(ws, col_idx: int, formula: str, alias_to_info: Dict,
                             start_row: int, output_file_path: Optional[str],
                             external_links: Optional[Dict], col_name: str) -> None:
    """应用公式到整列"""
    for row_idx in range(start_row, ws.max_row + 1):
        row_offset = row_idx - start_row
        adjusted_formula = replace_sheet_references(
            formula, alias_to_info, row_offset, output_file_path, external_links
        )
        ws.cell(row=row_idx, column=col_idx).value = adjusted_formula

    print(f"   已应用公式: {col_name} = {formula}")


# ==================== 主函数：生成Excel ====================

def generate_excel_from_template(
    template_file: str,
    template_sheet: str,
    formula_columns: List[str],
    data_sources: List[Dict],
    output_file: str,
    string_columns: Optional[List[str]] = None,
    use_external_refs: bool = False,
    primary_column: Optional[str] = None
) -> pd.DataFrame:
    """
    基于模板生成Excel文件

    Args:
        template_file: 模板文件路径
        template_sheet: 模板sheet名称
        formula_columns: 公式列集合
        data_sources: 数据源集合
        output_file: 输出文件路径
        string_columns: 字符串列列表
        use_external_refs: 是否使用外部引用公式
            - False（默认）: 将数据源sheet添加到输出文件，公式直接引用sheet名
            - True: 公式使用外部文件引用，不复制数据源sheet
        primary_column: 主键列名，为空时跳过过滤

    Returns:
        pd.DataFrame: 生成的数据
    """
    # 确保输出文件路径是绝对路径
    if not os.path.isabs(output_file):
        output_file = os.path.join(os.getcwd(), output_file)

    print("=== Excel模板生成器 ===\n")

    # 1. 验证文件
    _validate_input_files(template_file, data_sources)

    # 2. 分析模板
    external_links, template_formulas, template_columns, template_ws, template_wb = _analyze_template(
        template_file, template_sheet, formula_columns
    )

    # 3. 加载数据源
    loaded_data_sources, alias_to_info = _load_all_data_sources(
        data_sources, output_file, template_sheet, string_columns
    )

    # 4. 合并数据
    merged_df = merge_data_by_row(loaded_data_sources, template_columns)

    # 5. 过滤数据
    merged_df = _filter_data_by_primary_column(merged_df, primary_column)

    # 6. 生成输出文件
    _generate_output_file(
        output_file, merged_df, template_columns, template_formulas,
        formula_columns, alias_to_info, external_links, template_ws,
        use_external_refs, string_columns, data_sources
    )

    # 7. 打印公式汇总
    _print_formula_summary(output_file)

    # 关闭模板工作簿
    template_wb.close()

    return merged_df


def _validate_input_files(template_file: str, data_sources: List[Dict]) -> None:
    """验证输入文件是否存在"""
    print("1. 验证文件...")

    if not os.path.exists(template_file):
        raise FileNotFoundError(f"模板文件不存在: {template_file}")

    for ds in data_sources:
        if not os.path.exists(ds['file_path']):
            raise FileNotFoundError(f"数据文件不存在: {ds['file_path']}")

    print("   所有文件验证通过\n")


def _analyze_template(template_file: str, template_sheet: str,
                      formula_columns: List[str]) -> Tuple:
    """分析模板结构"""
    print("2. 分析模板...")

    external_links = read_external_links(template_file)
    template_columns, formula_templates, template_ws, template_wb = read_template_structure(
        template_file, template_sheet
    )

    # 收集模板中的公式
    template_formulas = {}
    for col_name in formula_columns:
        if col_name in formula_templates:
            template_formulas[col_name] = formula_templates[col_name]
        else:
            print(f"   警告: 公式列 '{col_name}' 在模板中没有公式")

    print()
    return external_links, template_formulas, template_columns, template_ws, template_wb


def _load_all_data_sources(data_sources: List[Dict], output_file: str,
                           template_sheet: str, string_columns: Optional[List[str]]) -> Tuple:
    """加载所有数据源"""
    print("3. 加载数据源...")

    loaded_data_sources = []
    alias_to_info = _init_alias_to_info(output_file, template_sheet)

    for idx, ds in enumerate(data_sources, start=1):
        config = _create_data_source_config(ds)
        df = read_data_source(config, string_columns)
        loaded_data_sources.append((config.alias, df))

        _update_alias_mapping(alias_to_info, str(idx), config)

    _print_data_source_mapping(output_file, data_sources)
    print()

    return loaded_data_sources, alias_to_info


def _init_alias_to_info(output_file: str, template_sheet: str) -> Dict:
    """初始化别名映射"""
    alias_to_info = {}

    # 输出 sheet 作为 0 号
    alias_to_info['0'] = {
        'file_path': output_file,
        'sheet_name': '结果'
    }

    # 模板 sheet 映射到输出 sheet
    alias_to_info[template_sheet.lower()] = {
        'file_path': output_file,
        'sheet_name': '结果',
        'is_template_self_reference': True
    }

    return alias_to_info


def _create_data_source_config(ds: Dict) -> DataSourceConfig:
    """创建数据源配置"""
    column_mappings = [
        DataColumnMapping(source_column=m['source'], target_column=m['target'])
        for m in ds['column_mappings']
    ]

    return DataSourceConfig(
        file_path=ds['file_path'],
        sheet_name=ds['sheet_name'],
        column_mappings=column_mappings,
        alias=ds.get('alias', '')
    )


def _update_alias_mapping(alias_to_info: Dict, idx: str, config: DataSourceConfig) -> None:
    """更新别名映射"""
    info = {
        'file_path': config.file_path,
        'sheet_name': config.sheet_name
    }

    alias_to_info[idx] = info

    if config.alias:
        alias_to_info[config.alias.lower()] = info

    alias_to_info[config.sheet_name.lower()] = info


def _print_data_source_mapping(output_file: str, data_sources: List[Dict]) -> None:
    """打印数据源映射"""
    print(f"   数据源编号映射:")
    print(f"     [0] -> 输出文件: {os.path.basename(output_file)} (sheet: 结果)")

    for idx, ds in enumerate(data_sources, start=1):
        print(f"     [{idx}] -> {os.path.basename(ds['file_path'])} (sheet: {ds['sheet_name']})")


def _filter_data_by_primary_column(merged_df: pd.DataFrame,
                                    primary_column: Optional[str]) -> pd.DataFrame:
    """根据主键列过滤数据"""
    if not primary_column:
        return merged_df

    if primary_column not in merged_df.columns:
        print(f"   警告: 主键列 '{primary_column}' 不存在于数据中，跳过过滤")
        return merged_df

    original_count = len(merged_df)
    merged_df = merged_df[
        merged_df[primary_column].notna() &
        (merged_df[primary_column].astype(str).str.strip() != '')
    ]

    filtered_count = len(merged_df)
    if original_count > filtered_count:
        print(f"   过滤掉 {original_count - filtered_count} 行（{primary_column} 列为空）")

    return merged_df


def _generate_output_file(output_file: str, merged_df: pd.DataFrame,
                          template_columns: List[str], template_formulas: Dict,
                          formula_columns: List[str], alias_to_info: Dict,
                          external_links: Dict, template_ws,
                          use_external_refs: bool, string_columns: Optional[List[str]],
                          data_sources: Optional[List[Dict]] = None) -> None:
    """生成输出文件"""
    print("5. 生成输出文件...")

    if use_external_refs:
        print("   模式: 外部引用公式（数据源保留在外部文件）")
    else:
        print("   模式: 内部引用公式（数据源sheet将复制到输出文件）")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df = merged_df.copy()

        # 准备输出DataFrame
        local_formula_columns = _prepare_output_df(
            output_df, template_formulas, formula_columns, use_external_refs
        )

        output_df.to_excel(writer, sheet_name='结果', index=False)
        ws = writer.sheets['结果']

        # 如果不使用外部引用，复制数据源sheet到输出文件
        if not use_external_refs and data_sources:
            _copy_data_source_sheets(writer, data_sources, alias_to_info)

        # 应用公式
        _apply_formulas(
            ws, output_df, template_formulas, formula_columns,
            alias_to_info, output_file, external_links, use_external_refs, local_formula_columns
        )

        # 应用模板样式
        apply_template_styles(ws, template_ws, template_columns, len(output_df) + 1)

        # 处理字符串列格式
        if string_columns:
            _apply_string_column_format(ws, output_df, string_columns)

    print(f"\n输出文件已保存: {output_file}")


def _copy_data_source_sheets(writer, data_sources: List[Dict], alias_to_info: Dict) -> None:
    """
    将数据源sheet复制到输出文件

    Args:
        writer: ExcelWriter对象
        data_sources: 数据源列表
        alias_to_info: 别名到信息的映射
    """
    print("   正在复制数据源sheet到输出文件...")

    copied_sheets = set()  # 跟踪已复制的sheet，避免重复

    for idx, ds in enumerate(data_sources, start=1):
        file_path = ds['file_path']
        sheet_name = ds['sheet_name']
        alias = ds.get('alias', '')

        # 确定目标sheet名称
        if alias:
            target_sheet_name = alias
        else:
            target_sheet_name = sheet_name

        # 避免sheet名冲突
        if target_sheet_name in copied_sheets or target_sheet_name == '结果':
            target_sheet_name = f"{target_sheet_name}_{idx}"

        # 读取数据源文件
        try:
            source_wb = openpyxl.load_workbook(file_path, data_only=False)
            if sheet_name not in source_wb.sheetnames:
                print(f"   警告: 数据源 {file_path} 中不存在sheet '{sheet_name}'")
                source_wb.close()
                continue

            source_ws = source_wb[sheet_name]

            # 创建新sheet
            target_ws = writer.book.create_sheet(title=target_sheet_name)

            # 复制所有单元格数据和样式
            _copy_worksheet(source_ws, target_ws)

            # 更新alias_to_info，使其指向内部的sheet
            info_update = {
                'file_path': '',  # 空路径表示内部引用
                'sheet_name': target_sheet_name,
                'is_internal': True
            }

            # 更新所有相关映射
            alias_to_info[str(idx)]['is_internal'] = True
            alias_to_info[str(idx)]['sheet_name'] = target_sheet_name
            alias_to_info[str(idx)]['file_path'] = ''

            if alias:
                if alias.lower() in alias_to_info:
                    alias_to_info[alias.lower()]['is_internal'] = True
                    alias_to_info[alias.lower()]['sheet_name'] = target_sheet_name
                    alias_to_info[alias.lower()]['file_path'] = ''

            if sheet_name.lower() in alias_to_info:
                alias_to_info[sheet_name.lower()]['is_internal'] = True
                alias_to_info[sheet_name.lower()]['sheet_name'] = target_sheet_name
                alias_to_info[sheet_name.lower()]['file_path'] = ''

            copied_sheets.add(target_sheet_name)
            print(f"   已复制: {os.path.basename(file_path)}[{sheet_name}] -> [{target_sheet_name}]")

            source_wb.close()
        except Exception as e:
            print(f"   警告: 无法复制数据源sheet '{sheet_name}': {e}")


def _copy_worksheet(source_ws, target_ws) -> None:
    """
    复制工作表的所有内容（数据和样式）

    Args:
        source_ws: 源工作表
        target_ws: 目标工作表
    """
    # 复制单元格数据和样式
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                copy_cell_style(cell, new_cell)

    # 复制合并单元格
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # 复制列宽
    for col_letter, col_dim in source_ws.column_dimensions.items():
        if col_dim.width:
            target_ws.column_dimensions[col_letter].width = col_dim.width

    # 复制行高
    for row_idx, row_dim in source_ws.row_dimensions.items():
        if row_dim.height:
            target_ws.row_dimensions[row_idx].height = row_dim.height


def _prepare_output_df(output_df: pd.DataFrame, template_formulas: Dict,
                        formula_columns: List[str], use_external_refs: bool) -> List[str]:
    """
    准备输出DataFrame

    Args:
        output_df: 输出DataFrame
        template_formulas: 模板公式字典
        formula_columns: 公式列列表
        use_external_refs: 是否使用外部引用

    Returns:
        List[str]: 需要应用公式的列名列表
    """
    # 收集所有需要应用公式的列
    applicable_formula_columns = []

    for col in formula_columns:
        if col in template_formulas:
            applicable_formula_columns.append(col)
            # 清空公式列，为后续写入公式做准备
            if col in output_df.columns:
                output_df[col] = None

    print(f"   公式列: {applicable_formula_columns}")

    return applicable_formula_columns


def _apply_formulas(ws, output_df: pd.DataFrame, template_formulas: Dict,
                    formula_columns: List[str], alias_to_info: Dict,
                    output_file: str, external_links: Dict,
                    use_external_refs: bool, formula_columns_to_apply: List[str]) -> None:
    """
    应用公式到工作表

    公式引用格式由 alias_to_info 中的 is_internal 标志决定：
    - is_internal=True: 内部引用（SheetName!A1）
    - is_internal=False: 外部引用（[filename]SheetName!A1）
    """
    if not formula_columns_to_apply or not template_formulas:
        return

    apply_formulas_to_output(
        ws, formula_columns_to_apply, template_formulas, alias_to_info,
        start_row=2, output_file_path=output_file, external_links=external_links
    )


def _find_column_index(df: pd.DataFrame, col_name: str) -> Optional[int]:
    """查找列索引"""
    for c_idx, c_name in enumerate(df.columns, start=1):
        if c_name == col_name:
            return c_idx
    return None


def _apply_string_column_format(ws, output_df: pd.DataFrame, string_columns: List[str]) -> None:
    """应用字符串列格式"""
    from openpyxl.styles import Font

    for col_idx, col_name in enumerate(output_df.columns, start=1):
        if col_name not in string_columns:
            continue

        ws.column_dimensions[get_column_letter(col_idx)].width = 15

        for row in range(2, len(output_df) + 2):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'

            if pd.notna(output_df.iloc[row - 2][col_name]):
                cell.value = str(output_df.iloc[row - 2][col_name])


def _print_formula_summary(output_file: str) -> None:
    """打印公式汇总"""
    print("\n" + "=" * 70)
    print("最终文件公式汇总")
    print("=" * 70)

    try:
        wb_output = openpyxl.load_workbook(output_file, data_only=False)
        ws_output = wb_output.active

        col_names = _read_output_column_names(ws_output)
        formulas = _read_output_formulas(ws_output, col_names)

        if formulas:
            print(f"\n公式列数: {len(formulas)}")
            for col_name, formula in formulas.items():
                formula_display = formula[:100] + "..." if len(formula) > 100 else formula
                print(f"  {col_name}: {formula_display}")
        else:
            print("\n无公式列（所有数据均为直接值）")

        wb_output.close()
    except Exception as e:
        print(f"   警告: 无法读取输出文件公式: {e}")

    print("=" * 70)


def _read_output_column_names(ws) -> Dict[int, str]:
    """读取输出文件的列名"""
    col_names = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            col_names[col_idx] = str(cell.value)
    return col_names


def _read_output_formulas(ws, col_names: Dict[int, str]) -> Dict[str, str]:
    """读取输出文件的公式"""
    formulas = {}

    if ws.max_row < 2:
        return formulas

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            col_name = col_names.get(col_idx, f"列{col_idx}")
            formulas[col_name] = cell.value

    return formulas


# ==================== 命令行解析辅助函数 ====================

def parse_column_mappings(mappings_str: str) -> List[Dict[str, str]]:
    """
    解析列映射字符串

    Args:
        mappings_str: 列映射字符串，格式为 "SourceCol:TargetCol,SourceCol2:TargetCol2"
                      或 "Col1,Col2"（源列名和目标列名相同）

    Returns:
        List[Dict[str, str]]: 列映射列表
    """
    mappings = []

    if not mappings_str:
        return mappings

    pairs = mappings_str.split(',')

    for pair in pairs:
        if ':' in pair:
            source, target = pair.split(':', 1)
            mappings.append({
                'source': source.strip(),
                'target': target.strip()
            })
        else:
            # 不带冒号时，源列名和目标列名相同
            col_name = pair.strip()
            if col_name:
                mappings.append({
                    'source': col_name,
                    'target': col_name
                })

    return mappings
