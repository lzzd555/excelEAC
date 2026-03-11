"""
Excel数据验证模块
提供数据验证、分组和异常检测功能
"""

import pandas as pd
from typing import List, Dict, Any, Optional
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def process_excel_with_validation(
    input_file: str,
    sheet_name: str,
    group_columns: List[str],
    compare_columns: List[str],
    output_columns: Optional[List[str]] = None,
    output_file: str = 'validation_result.xlsx',
    string_columns: Optional[List[str]] = None,
    abnormal_detail_columns: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    处理Excel文件并进行数据验证

    参数:
        input_file: 输入Excel文件路径
        sheet_name: 工作表名称
        group_columns: 分组列名列表
        compare_columns: 需要比较是否相等的列名列表（必须是2列）
        output_columns: 输出到新Excel的列名列表
        output_file: 输出文件名
        string_columns: 需要保持为字符串格式的列名列表（避免"001"变成1）
        abnormal_detail_columns: 异常详情中需要显示的原表列名列表。如果为None，则自动包含分组列、比较列和字符串列。

    返回:
        处理后的DataFrame
    """
    # 确保输出文件在当前文件夹
    if not os.path.isabs(output_file):
        # 如果是相对路径，确保输出到当前文件夹
        output_file = os.path.join(os.getcwd(), output_file)

    # 1. 读取Excel数据，保持原始格式
    print(f"正在读取 {input_file} 的 {sheet_name} 工作表...")

    # 构建dtype字典，保持字符串格式
    dtype_dict = {}
    if string_columns:
        for col in string_columns:
            dtype_dict[col] = 'string'

    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=dtype_dict)

    # 2. 数据验证检查
    print("正在验证数据...")
    # 获取比较列名
    col1, col2 = compare_columns[0], compare_columns[1]

    # 检查数据是否有效
    if len(compare_columns) != 2:
        raise ValueError("compare_columns 必须包含 exactly 2 列名")

    if col1 not in df.columns or col2 not in df.columns:
        missing_cols = [col for col in compare_columns if col not in df.columns]
        raise ValueError(f"比较列不存在: {missing_cols}")

    # 确保字符串列保持字符串格式
    if string_columns:
        for col in string_columns:
            if col in df.columns:
                # 使用pandas的string类型来保持前导零
                df[col] = df[col].astype('string')

    # 逐行比较两列数据
    df['行是否正常'] = (df[col1] == df[col2])

    # 3. 分组检查
    print("正在进行分组验证...")

    # 使用更直接的方法计算统计信息
    # 获取所有唯一的分组
    unique_groups = df.groupby(group_columns).size().reset_index()

    # 计算每个组的状态
    group_results = []
    for _, group_row in unique_groups.iterrows():
        group_key = tuple(group_row[group_columns])

        # 筛选出属于当前组的数据
        group_data = df.copy()
        for i, col in enumerate(group_columns):
            group_data = group_data[group_data[col] == group_key[i]]

        # 检查是否全正常
        is_all_normal = group_data['行是否正常'].all()
        normal_count = group_data['行是否正常'].sum()
        total_count = len(group_data)
        abnormal_count = total_count - normal_count

        # 创建结果行
        result_row = group_row.to_dict()
        result_row['验证状态'] = '正常' if is_all_normal else '异常'
        result_row['正常行数'] = normal_count
        result_row['异常行数'] = abnormal_count
        result_row['总行数'] = total_count
        result_row['异常率'] = f"{abnormal_count/total_count*100:.1f}%"

        group_results.append(result_row)

    # 创建结果DataFrame
    group_stats = pd.DataFrame(group_results)

    # 4. 处理输出列选择
    # 确保验证状态在输出列中
    if output_columns is not None:
        # 只保留分组列（因为分组后的结果不包含原始数据列）
        final_output_columns = group_columns.copy()
        # 添加验证状态列
        if '验证状态' not in final_output_columns:
            final_output_columns.append('验证状态')
    else:
        # 如果没有指定输出列，使用所有列
        final_output_columns = group_stats.columns.tolist()

    # 5. 选择最终的组级别数据（只保留汇总数据）
    final_result = group_stats[final_output_columns].copy()

    # 6. 如果用户指定了额外的输出列，尝试添加汇总信息
    # 注意：只有已经在 group_stats 中的列才能被添加
    if output_columns is not None:
        # 检查是否有在output_columns中但不在final_output_columns中的列
        additional_cols = [col for col in output_columns
                          if col in group_stats.columns and col not in final_output_columns]
        if additional_cols:
            # 添加这些汇总列
            for col in additional_cols:
                if col in group_stats.columns:
                    final_result[col] = group_stats[col]
        else:
            # 如果没有额外的列可用，显示可用的列
            print(f"\n可用的汇总列: {list(group_stats.columns)}")
            print(f"注意: 不能在汇总数据中包含原始数据列（如订单号、产品代码）")
            print(f"因为这些列在分组后不存在于汇总结果中\n")

    # 6. 处理字符串格式列
    if string_columns:
        # 确保分组列中的字符串列保持格式
        for col in group_columns:
            if col in string_columns and col in final_result.columns:
                final_result[col] = final_result[col].astype(str)

    # 7. 输出到Excel，保持原始格式
    print(f"正在输出结果到 {output_file}...")

    # 创建Excel writer，使用openpyxl保持格式
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 验证结果 - 组级别的汇总数据
        # 创建副本并强制转换为字符串格式
        final_result_copy = final_result.copy()

        if string_columns:
            for col in string_columns:
                if col in final_result_copy.columns:
                    final_result_copy[col] = final_result_copy[col].astype(str)

        final_result_copy.to_excel(writer, sheet_name='验证结果', index=False)

        # 分组统计摘要 - 更详细的统计信息
        group_summary = group_stats[group_columns + ['验证状态', '正常行数', '异常行数', '总行数', '异常率']]
        group_summary_copy = group_summary.copy()

        # 同样处理分组统计中的字符串列
        if string_columns:
            for col in string_columns:
                if col in group_summary_copy.columns:
                    group_summary_copy[col] = group_summary_copy[col].astype(str)

        group_summary_copy.to_excel(writer, sheet_name='分组统计', index=False)

        # 异常详情（现在包含所有行，标记异常情况）
        if '行是否正常' in df.columns:
            # 包含所有数据行，但标记异常情况
            detail_data = df.copy()

            # 添加异常状态标记列
            detail_data['是否异常'] = ~detail_data['行是否正常']

            # 构建异常详情列列表
            # 首先添加分组列
            abnormal_cols = group_columns.copy()

            # 如果指定了异常详情列，则使用指定的列
            if abnormal_detail_columns:
                for col in abnormal_detail_columns:
                    if col in detail_data.columns and col not in abnormal_cols:
                        abnormal_cols.append(col)
            else:
                # 如果没有指定，自动包含重要的列
                # 添加字符串列
                if string_columns:
                    for col in string_columns:
                        if col in detail_data.columns and col not in abnormal_cols:
                            abnormal_cols.append(col)

                # 添加比较列
                for col in compare_columns:
                    if col not in abnormal_cols:
                        abnormal_cols.append(col)

            # 添加状态列
            abnormal_cols.extend(['行是否正常', '是否异常'])

            # 只保留在原始数据中存在的列
            valid_cols = [col for col in abnormal_cols if col in detail_data.columns]
            detail_data_copy = detail_data[valid_cols].copy()

            # 应用字符串格式 - 使用更可靠的方法保持字符串格式
            if string_columns:
                for col in string_columns:
                    if col in detail_data_copy.columns:
                        # 确保使用原始数据中的正确字符串值
                        original_values = df[col].tolist()
                        for i in range(len(detail_data_copy)):
                            # 从原始DataFrame获取值，确保前导零不被丢失
                            original_index = df.index[i] if i < len(df) else None
                            if original_index is not None and col in df.loc[original_index]:
                                original_value = df.loc[original_index, col]
                                if pd.notna(original_value):
                                    detail_data_copy.iloc[i, detail_data_copy.columns.get_loc(col)] = str(original_value)
                        # 转换为string类型以保持格式
                        detail_data_copy[col] = detail_data_copy[col].astype('string')

            # 直接使用处理好的数据创建临时DataFrame
            temp_abnormal_data = detail_data_copy.copy()

            # 使用保存好格式的数据写入Excel
            temp_abnormal_data.to_excel(writer, sheet_name='异常详情', index=False)

            # 获取工作表进行额外配置
            ws = writer.sheets['异常详情']

            # 为字符串列设置格式，确保前导零不被去掉
            if string_columns:
                for col_idx, col_name in enumerate(temp_abnormal_data.columns):
                    if col_name in string_columns:
                        # 设置列宽，确保内容显示完整
                        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 15

                        # 设置整个列的格式为文本
                        for row in range(1, len(temp_abnormal_data) + 2):  # 包含标题行
                            cell = ws.cell(row=row, column=col_idx + 1)
                            # 设置单元格格式为文本，保持前导零
                            cell.number_format = '@'

                        # 确保数据是字符串格式
                        for row_idx in range(len(temp_abnormal_data)):
                            row_idx_excel = row_idx + 2  # Excel行号（从2开始）
                            cell_value = temp_abnormal_data.iloc[row_idx, col_idx]
                            if pd.notna(cell_value):
                                # 直接设置字符串值
                                ws.cell(row=row_idx_excel, column=col_idx + 1).value = str(cell_value)

            # 添加颜色标记：异常行标红
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # 为异常行添加背景色
            for row_idx in range(len(temp_abnormal_data)):
                row_idx_excel = row_idx + 2  # Excel行号（从2开始）
                is_abnormal = temp_abnormal_data.iloc[row_idx, temp_abnormal_data.columns.get_loc('是否异常')]
                if is_abnormal:
                    # 为整行添加红色背景
                    for col_idx in range(len(temp_abnormal_data.columns)):
                        cell = ws.cell(row=row_idx_excel, column=col_idx + 1)
                        cell.fill = red_fill

    print(f"文件已保存到: {output_file}")
    return final_result
