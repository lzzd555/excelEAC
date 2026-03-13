"""
调试填充样式复制问题
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 设置各种填充样式
# 1. 无填充
ws.cell(row=1, column=1).value = "No Fill"
ws.cell(row=1, column=1).fill = PatternFill(fill_type='none')

# 2. RGB填充
ws.cell(row=1, column=2).value = "RGB Fill"
ws.cell(row=1, column=2).fill = PatternFill(fill_type='solid', start_color='FFCC99')

# 3. 索引颜色填充
ws.cell(row=1, column=3).value = "Indexed Fill"
ws.cell(row=1, column=3).fill = PatternFill(
    fill_type='solid',
    start_color=openpyxl.styles.colors.Color(indexed=15)
)

# 4. 灰度填充
ws.cell(row=1, column=4).value = "Gray 125"
ws.cell(row=1, column=4).fill = PatternFill(fill_type='gray125')

ws.cell(row=1, column=5).value = "Gray 0625"
ws.cell(row=1, column=5).fill = PatternFill(fill_type='gray0625')

# 保存原始文件
wb.save('debug_original.xlsx')
wb.close()

# 重新加载并复制样式
wb_orig = openpyxl.load_workbook('debug_original.xlsx', data_only=False)
ws_orig = wb_orig.active

wb_copy = openpyxl.Workbook()
ws_copy = wb_copy.active

print("原始文件样式:")
for col in range(1, 6):
    cell = ws_orig.cell(row=1, column=col)
    print(f"列 {col}: 类型={cell.fill.fill_type if cell.fill else 'None'}, 颜色={cell.fill.start_color if cell.fill else 'None'}")

print("\n开始复制样式...")
for col in range(1, 6):
    src_cell = ws_orig.cell(row=1, column=col)
    dst_cell = ws_copy.cell(row=1, column=col)
    dst_cell.value = src_cell.value

    if src_cell.has_style:
        print(f"\n复制列 {col} 的样式...")
        print(f"  原始填充类型: {src_cell.fill.fill_type if src_cell.fill else 'None'}")

        if src_cell.fill:
            # 尝试复制填充
            try:
                # 方法1: 直接复制填充对象
                dst_cell.fill = src_cell.fill
                print(f"  ✓ 方法1成功: 类型={dst_cell.fill.fill_type if dst_cell.fill else 'None'}")
            except Exception as e:
                print(f"  ✗ 方法1失败: {e}")
                try:
                    # 方法2: 重建填充
                    if src_cell.fill.fill_type == 'solid':
                        dst_cell.fill = PatternFill(
                            fill_type='solid',
                            start_color=src_cell.fill.start_color.rgb if hasattr(src_cell.fill.start_color, 'rgb') else 'FFFFFF'
                        )
                        print(f"  ✓ 方法2成功: 类型={dst_cell.fill.fill_type if dst_cell.fill else 'None'}")
                    else:
                        dst_cell.fill = PatternFill(fill_type=src_cell.fill.fill_type)
                        print(f"  ✓ 方法2成功: 类型={dst_cell.fill.fill_type if dst_cell.fill else 'None'}")
                except Exception as e2:
                    print(f"  ✗ 方法2失败: {e2}")
                    print(f"  跳过填充复制")

wb_copy.save('debug_copy.xlsx')
wb_orig.close()

# 验证结果
wb_result = openpyxl.load_workbook('debug_copy.xlsx', data_only=False)
ws_result = wb_result.active

print("\n复制结果:")
for col in range(1, 6):
    cell = ws_result.cell(row=1, column=col)
    print(f"列 {col}: 类型={cell.fill.fill_type if cell.fill else 'None'}, 颜色={cell.fill.start_color if cell.fill else 'None'}")

wb_result.close()