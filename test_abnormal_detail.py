#!/usr/bin/env python3
"""
专门测试异常详情中的 string_columns 效果
"""

import pandas as pd
from excel_validator import process_excel_with_validation

# 创建有异常的数据
print("=== 测试异常详情中的 string_columns ===\n")

data = {
    '订单号': ['001', '002', '003', '004'],
    '产品代码': ['001', '002', '001', '003'],
    '部门': ['A', 'A', 'B', 'B'],
    '计划值': [100, 200, 150, 120],
    '实际值': [100, 210, 150, 120]  # ORD002 有异常
}

df = pd.DataFrame(data)
df.to_excel('abnormal_test.xlsx', index=False)

print("测试数据：")
print(df)

print("\n运行验证...")
result = process_excel_with_validation(
    input_file='abnormal_test.xlsx',
    sheet_name='Sheet1',
    group_columns=['部门'],
    compare_columns=['计划值', '实际值'],
    output_columns=['部门', '验证状态', '总行数'],
    output_file='test_output.xlsx',
    string_columns=['订单号', '产品代码']  # 应该包含在异常详情中
)

print("\n分组结果：")
print(result)

print("\n检查异常详情...")
try:
    # 直接读取Excel文件
    import openpyxl
    wb = openpyxl.load_workbook('test_output.xlsx')
    if '异常详情' in wb.sheetnames:
        detail_sheet = wb['异常详情']
        print("\n异常详情内容：")
        for row in detail_sheet.iter_rows(values_only=True):
            print(row)

        # 检查订单号格式
        print("\n检查订单号列的数据：")
        order_ids = []
        for row in detail_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 非空行
                order_ids.append(row[0])
        print(f"订单号: {order_ids}")
        print(f"订单号类型: {[type(id) for id in order_ids]}")
    else:
        print("❌ 没有异常详情sheet")

except Exception as e:
    print(f"读取失败: {e}")

# 清理
import os
# os.remove('abnormal_test.xlsx')
# os.remove('test_output.xlsx')

print("\n测试完成！")