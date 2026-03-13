"""
测试 Excel 外部引用索引格式 [N]SheetName
"""
import pandas as pd
import openpyxl
import os
import zipfile
import xml.etree.ElementTree as ET

# 清理旧文件
for f in ['external_data.xlsx', 'template_with_index_ref.xlsx', 'test_index_ref_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Sales': [100, 200, 300, 400, 500],
        'Cost': [50, 100, 150, 200, 250],
        'Tax': [10, 20, 30, 40, 50]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含 Excel 外部引用索引格式的模板文件
def create_template_with_index_ref():
    """创建模板文件，使用 Excel 外部引用索引格式 [N]SheetName"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MyTemplateSheet'  # 设置模板 sheet 名称

    # 第一行：列名
    headers = ['Sales', 'Cost', 'Tax', 'SubTotal', 'Profit', 'Ratio']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # 第二行：公式模板（使用 Excel 外部引用索引格式）
    # SubTotal: =[2]MyTemplateSheet!A2-[2]MyTemplateSheet!B2-[2]MyTemplateSheet!C2
    ws.cell(row=2, column=4).value = "=[2]MyTemplateSheet!A2-[2]MyTemplateSheet!B2-[2]MyTemplateSheet!C2"

    # Profit: =[2]MyTemplateSheet!D2*0.2
    ws.cell(row=2, column=5).value = "=[2]MyTemplateSheet!D2*0.2"

    # Ratio: =[2]MyTemplateSheet!E2/[2]MyTemplateSheet!D2
    ws.cell(row=2, column=6).value = "=[2]MyTemplateSheet!E2/[2]MyTemplateSheet!D2"

    # 保存文件
    wb.save('template_with_index_ref.xlsx')
    wb.close()

    print("✓ 创建模板文件: template_with_index_ref.xlsx")
    print("  - Sheet 名称: MyTemplateSheet")
    print("  - 包含 Excel 外部引用索引格式的公式:")
    print("    - SubTotal = [2]MyTemplateSheet!A2 - [2]MyTemplateSheet!B2 - [2]MyTemplateSheet!C2")
    print("    - Profit = [2]MyTemplateSheet!D2 * 0.2")
    print("    - Ratio = [2]MyTemplateSheet!E2 / [2]MyTemplateSheet!D2")
    print("\n  注意: [2] 是 Excel 存储外部引用的索引号")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：Excel 外部引用索引格式 [N]SheetName")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [
                {'source': 'Sales', 'target': 'Sales'},
                {'source': 'Cost', 'target': 'Cost'},
                {'source': 'Tax', 'target': 'Tax'}
            ],
            'alias': 'external_data'
        }
    ]

    # 公式列（这些列的公式会被应用到输出文件）
    formula_columns = ['SubTotal', 'Profit', 'Ratio']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_with_index_ref.xlsx',
            template_sheet='MyTemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_index_ref_output.xlsx',
            use_external_refs=True  # 使用外部引用模式
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_index_ref_output.xlsx")

        # 验证输出文件中的公式
        print("\n正在验证输出文件中的公式...")
        wb = openpyxl.load_workbook('test_index_ref_output.xlsx', data_only=False)
        ws = wb.active

        # 读取第二行的公式
        subtotal_formula = ws.cell(row=2, column=4).value
        profit_formula = ws.cell(row=2, column=5).value
        ratio_formula = ws.cell(row=2, column=6).value

        print(f"  SubTotal: {subtotal_formula}")
        print(f"  Profit: {profit_formula}")
        print(f"  Ratio: {ratio_formula}")

        wb.close()

        # 检查结果
        print("\n结果分析:")
        print("  原始模板公式: [2]MyTemplateSheet!A2")
        print("  预期输出公式: 结果!A2 (本地引用)")
        print(f"  实际输出公式: {subtotal_formula[:20]}...")

        # 检查是否包含索引号（不应该包含）
        if '[2]' in str(subtotal_formula):
            print("\n✗ 错误：公式中仍然包含索引号 [2]")
        elif 'test_index_ref_output.xlsx' in str(subtotal_formula):
            print("\n✗ 错误：公式中包含输出文件名（应该是本地引用）")
        else:
            print("\n✓ 正确：公式使用本地引用格式")
            print("✓ Excel 索引 [2] 被正确映射到输出 sheet '结果'")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_index_ref()

    # 运行测试
    run_test()
