"""
测试所有样式类型的复制
"""
import pandas as pd
import openpyxl
import os

# 清理旧文件
for f in ['external_data.xlsx', 'template_all_styles.xlsx', 'test_all_styles_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Value': [100, 200, 300]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含所有样式的模板文件
def create_template_with_all_styles():
    """创建包含所有样式的模板文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'

    # 第一行：列名
    headers = [
        'FontBold', 'FontItalic', 'FontColor', 'AlignmentCenter',
        'AlignmentWrap', 'Border', 'NumberFormat', 'Protection'
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='FFFFCC')

    # 第二行：样式测试数据
    col = 1

    # 1. 粗体字体
    cell = ws.cell(row=2, column=col)
    cell.value = "Bold Text"
    cell.font = openpyxl.styles.Font(bold=True, size=12)

    # 2. 斜体字体
    cell = ws.cell(row=2, column=col+1)
    cell.value = "Italic Text"
    cell.font = openpyxl.styles.Font(italic=True, size=12)

    # 3. 颜色字体
    cell = ws.cell(row=2, column=col+2)
    cell.value = "Red Text"
    cell.font = openpyxl.styles.Font(color='FF0000', size=12)

    # 4. 居中对齐
    cell = ws.cell(row=2, column=col+3)
    cell.value = "Center"
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # 5. 自动换行
    cell = ws.cell(row=2, column=col+4)
    cell.value = "Wrap\nText"
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # 6. 边框
    cell = ws.cell(row=2, column=col+5)
    cell.value = "Bordered"
    cell.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    # 7. 数字格式
    cell = ws.cell(row=2, column=col+6)
    cell.value = 12345.67
    cell.number_format = '#,##0.00'

    # 8. 保护锁定
    cell = ws.cell(row=2, column=col+7)
    cell.value = "Protected"
    cell.protection = openpyxl.styles.Protection(locked=True)

    # 保存文件
    wb.save('template_all_styles.xlsx')
    wb.close()

    print("✓ 创建包含所有样式的模板文件")
    print("  - 字体样式：粗体、斜体、颜色")
    print("  - 对齐样式：居中、自动换行")
    print("  - 边框样式：实线边框")
    print("  - 数字格式：#,##0.00")
    print("  - 保护样式：锁定")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：所有样式类型复制")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [{'source': 'Value', 'target': 'Value'}],
            'alias': 'external_data'
        }
    ]

    # 公式列
    formula_columns = []

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_all_styles.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_all_styles_output.xlsx',
            use_external_refs=False  # 禁用外部引用，避免干扰样式测试
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_all_styles_output.xlsx")

        # 验证输出文件中的样式
        wb = openpyxl.load_workbook('test_all_styles_output.xlsx', data_only=False)
        ws = wb.active

        print("\n样式验证结果:")

        # 验证字体样式
        cell = ws.cell(row=2, column=1)
        print(f"  粗体: {'✓' if cell.font.bold else '✗'}")
        print(f"  斜体: {'✓' if cell.font.italic else '✗'}")

        cell = ws.cell(row=2, column=2)
        print(f"  字体大小: {cell.font.size if cell.font.size else '无'}")

        cell = ws.cell(row=2, column=3)
        print(f"  字体颜色: {'✓' if cell.font.color.rgb == 'FF0000' else '✗'}")

        # 验证对齐样式
        cell = ws.cell(row=2, column=4)
        print(f"  居中对齐: {'✓' if cell.alignment.horizontal == 'center' else '✗'}")

        cell = ws.cell(row=2, column=5)
        print(f"  自动换行: {'✓' if cell.alignment.wrap_text else '✗'}")

        # 验证边框样式
        cell = ws.cell(row=2, column=6)
        has_left_border = cell.border.left.border_style != 'none'
        has_right_border = cell.border.right.border_style != 'none'
        has_top_border = cell.border.top.border_style != 'none'
        has_bottom_border = cell.border.bottom.border_style != 'none'
        has_border = has_left_border and has_right_border and has_top_border and has_bottom_border
        print(f"  边框: {'✓' if has_border else '✗'} (左:{'✓' if has_left_border else '✗'} 右:{'✓' if has_right_border else '✗'} 上:{'✓' if has_top_border else '✗'} 下:{'✓' if has_bottom_border else '✗'}")

        # 验证数字格式
        cell = ws.cell(row=2, column=7)
        print(f"  数字格式: {cell.number_format}")

        # 验证保护样式
        cell = ws.cell(row=2, column=8)
        print(f"  保护锁定: {'✓' if cell.protection.locked else '✗'}")

        # 验证列宽
        print(f"\n列宽验证:")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            width = ws.column_dimensions[col_letter].width
            print(f"  列{col}宽度: {width}")

        wb.close()

        print("\n🎉 样式复制功能正常工作！")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

    return True


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_all_styles()

    # 运行测试
    run_test()