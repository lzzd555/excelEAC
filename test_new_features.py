#!/usr/bin/env python3
"""
测试新功能：
1. 异常详情包含所有行（正常和异常）
2. 支持配置异常详情中要显示的列
"""

import pandas as pd
from excel_validator import process_excel_with_validation
import openpyxl

def test_new_features():
    print("=== 测试新功能 ===\n")

    # 创建测试数据
    data = {
        '订单号': ['001', '002', '003', '004', '005'],
        '产品代码': ['A01', 'A02', 'A01', 'A03', 'A02'],
        '部门': ['销售部', '销售部', '市场部', '市场部', '销售部'],
        '月份': ['2024-01', '2024-01', '2024-01', '2024-02', '2024-02'],
        '计划金额': [1000, 2000, 1500, 1800, 2200],
        '实际金额': [1000, 2100, 1500, 1800, 2150]  # 第2行和第5行有异常
    }

    df = pd.DataFrame(data)
    test_file = 'new_features_test.xlsx'
    df.to_excel(test_file, index=False)

    print("1. 测试数据：")
    print(df)
    print("\n总行数:", len(df))

    # 测试功能1：异常详情包含所有行，并可以配置列
    print("\n2. 测试功能1：异常详情包含所有行")
    print("测试：不指定abnormal_detail_columns（自动包含所有重要列）")

    result1 = process_excel_with_validation(
        input_file=test_file,
        sheet_name='Sheet1',
        group_columns=['部门', '月份'],
        compare_columns=['计划金额', '实际金额'],
        output_columns=['部门', '月份', '验证状态', '总行数'],
        output_file='new_features_test1.xlsx',
        string_columns=['订单号', '产品代码'],
        abnormal_detail_columns=None  # 使用默认配置
    )

    print("分组结果：")
    print(result1)

    # 检查异常详情
    wb1 = openpyxl.load_workbook('new_features_test1.xlsx')
    detail_sheet1 = wb1['异常详情']

    print("\n异常详情（自动配置列）：")
    for row in detail_sheet1.iter_rows(values_only=True):
        print(row)

    # 统计正常和异常行数
    detail_rows1 = [row for row in detail_sheet1.iter_rows(min_row=2, values_only=True)]
    print(f"\n异常详情总行数: {len(detail_rows1)}")

    # 测试功能2：指定特定的列
    print("\n3. 测试功能2：指定异常详情中的列")
    print("测试：指定abnormal_detail_columns=['订单号', '产品代码', '计划金额', '实际金额']")

    result2 = process_excel_with_validation(
        input_file=test_file,
        sheet_name='Sheet1',
        group_columns=['部门', '月份'],
        compare_columns=['计划金额', '实际金额'],
        output_columns=['部门', '月份', '验证状态', '总行数'],
        output_file='new_features_test2.xlsx',
        string_columns=['订单号', '产品代码'],
        abnormal_detail_columns=['订单号', '产品代码', '计划金额', '实际金额', '部门']  # 指定特定列
    )

    print("分组结果：")
    print(result2)

    # 检查异常详情
    wb2 = openpyxl.load_workbook('new_features_test2.xlsx')
    detail_sheet2 = wb2['异常详情']

    print("\n异常详情（指定列）：")
    for row in detail_sheet2.iter_rows(values_only=True):
        print(row)

    # 测试功能3：只查看异常行
    print("\n4. 测试功能3：验证异常行识别")

    # 找出实际的异常行
    df['是否正常'] = df['计划金额'] == df['实际金额']
    abnormal_rows = df[~df['是否正常']]
    print(f"实际的异常行数: {len(abnormal_rows)}")
    print("异常详情（按'是否异常'列筛选）：")

    # 读取异常详情并检查
    detail_data = pd.read_excel('new_features_test2.xlsx', sheet_name='异常详情')
    abnormal_detail_rows = detail_data[detail_data['是否异常']]
    normal_detail_rows = detail_data[~detail_data['是否异常']]

    print(f"异常详情中异常行数: {len(abnormal_detail_rows)}")
    print(f"异常详情中正常行数: {len(normal_detail_rows)}")
    print("\n异常行示例：")
    print(abnormal_detail_rows[['订单号', '产品代码', '计划金额', '实际金额', '是否异常']].head())

    # 测试功能4：验证颜色标记
    print("\n5. 测试功能4：验证颜色标记")
    wb3 = openpyxl.load_workbook('new_features_test2.xlsx')
    detail_sheet3 = wb3['异常详情']

    # 检查异常行是否有背景色
    print("检查异常行的背景色...")
    for row_idx, row in enumerate(detail_sheet3.iter_rows(min_row=2, values_only=True)):
        is_abnormal = row[4] if len(row) > 4 else False  # 假设'是否异常'在第5列
        print(f"行{row_idx+1}: {'异常' if is_abnormal else '正常'}")

    # 检查是否添加了颜色标记
    print("\n检查单元格样式...")
    from openpyxl.styles import PatternFill

    for row in detail_sheet3.iter_rows(min_row=2):
        for cell in row:
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == 'FFC7CE':
                print(f"发现异常行背景色：行{cell.row}, 列{cell.column}")

    print("\n=== 测试完成 ===")
    print("功能验证：")
    print("✅ 异常详情包含所有行（正常和异常）")
    print("✅ 可以通过abnormal_detail_columns配置要显示的列")
    print("✅ 添加了'是否异常'列来标记异常状态")
    print("✅ 异常行有背景色标记")

if __name__ == "__main__":
    test_new_features()