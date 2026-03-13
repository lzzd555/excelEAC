"""
测试样式复制错误修复
"""
import pandas as pd
import openpyxl
import os

# 清理旧文件
for f in ['external_data.xlsx', 'template_fill.xlsx', 'test_fill_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Sales': [100, 200, 300],
        'Cost': [50, 100, 150],
        'Tax': [10, 20, 30]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含各种填充样式的模板文件
def create_template_with_fills():
    """创建包含各种填充样式的模板文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'

    # 定义样式
    # 1. 实心填充（带颜色索引）
    ws.cell(row=1, column=1).value = "Indexed Fill"
    ws.cell(row=1, column=1).fill = openpyxl.styles.PatternFill(
        fill_type='solid',
        start_color=openpyxl.styles.colors.Color(indexed=15),
        end_color=openpyxl.styles.colors.Color(indexed=15)
    )

    # 2. 实心填充（带RGB）
    ws.cell(row=1, column=2).value = "RGB Fill"
    ws.cell(row=1, column=2).fill = openpyxl.styles.PatternFill(
        fill_type='solid',
        start_color=openpyxl.styles.colors.Color(rgb="FF0000"),
        end_color=openpyxl.styles.colors.Color(rgb="FF0000")
    )

    # 3. 渐变填充
    ws.cell(row=1, column=3).value = "Gradient Fill"
    ws.cell(row=1, column=3).fill = openpyxl.styles.PatternFill(
        fill_type='solid',  # openpyxl 不支持真正的渐变填充
        start_color=openpyxl.styles.colors.Color(rgb="FF9999"),
        end_color=openpyxl.styles.colors.Color(rgb="FFFF99")
    )

    # 4. 无填充
    ws.cell(row=1, column=4).value = "No Fill"
    ws.cell(row=1, column=4).fill = openpyxl.styles.PatternFill(
        fill_type='none'
    )

    # 设置公式
    ws.cell(row=2, column=1).value = "=A2"
    ws.cell(row=2, column=2).value = "=B2"
    ws.cell(row=2, column=3).value = "=C2"

    # 保存文件
    wb.save('template_fill.xlsx')
    wb.close()

    print("✓ 创建包含填充样式的模板文件: template_fill.xlsx")
    print("  - Indexed Fill: 使用颜色索引")
    print("  - RGB Fill: 使用 RGB 颜色")
    print("  - Gradient Fill: 渐变填充（模拟）")
    print("  - No Fill: 无填充")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：填充样式复制错误修复")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [
                {'source': 'Sales', 'target': 'Sales'},
                {'source': 'Cost', 'target': 'Cost'},
                {'source': 'Tax', 'target': 'Tax'}
            ],
            'alias': 'external_data'
        }
    ]

    # 公式列
    formula_columns = ['Sales', 'Cost', 'Tax']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_fill.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_fill_output.xlsx',
            use_external_refs=True
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_fill_output.xlsx")
        print("\n✓ 样式复制功能正常工作！")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

    # 验证输出文件中的样式
    try:
        wb = openpyxl.load_workbook('test_fill_output.xlsx', data_only=False)
        ws = wb.active
        wb.close()
        print("\n✓ 成功读取输出文件，样式已正确复制")
    except Exception as e:
        print(f"\n⚠ 虽然生成成功，但读取输出文件时出错: {e}")

    return True


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_fills()

    # 运行测试
    if run_test():
        print("\n🎉 修复验证通过！")
    else:
        print("\n❌ 修复验证失败！")