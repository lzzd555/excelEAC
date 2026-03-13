"""
测试样式保留功能
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# 清理旧文件
for f in ['external_data.xlsx', 'template_with_styles.xlsx', 'test_styles_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Sales': [100, 200, 300, 400, 500],
        'Cost': [50, 100, 150, 200, 250],
        'Tax': [10, 20, 30, 40, 50]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建带样式的模板文件
def create_template_with_styles():
    """创建带样式的模板文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'

    # 定义样式
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    data_font = Font(name='Arial', size=10)
    data_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    data_alignment = Alignment(horizontal='right', vertical='center')

    formula_font = Font(name='Arial', size=10, bold=True, color='FF0000')
    formula_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')

    # 第一行：列名（带样式）
    headers = ['Sales', 'Cost', 'Tax', 'SubTotal', 'Profit', 'Ratio']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 第二行：公式模板（带样式）
    # Sales, Cost, Tax 列：数据列样式
    for col_idx in range(1, 4):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = data_font
        cell.fill = data_fill
        cell.alignment = data_alignment

    # SubTotal 列：公式样式
    ws.cell(row=2, column=4).value = "=A2-B2-C2"
    ws.cell(row=2, column=4).font = formula_font
    ws.cell(row=2, column=4).fill = formula_fill
    ws.cell(row=2, column=4).alignment = Alignment(horizontal='right', vertical='center')

    # Profit 列：公式样式
    ws.cell(row=2, column=5).value = "=D2*0.2"
    ws.cell(row=2, column=5).font = formula_font
    ws.cell(row=2, column=5).fill = formula_fill
    ws.cell(row=2, column=5).alignment = Alignment(horizontal='right', vertical='center')

    # Ratio 列：公式样式
    ws.cell(row=2, column=6).value = "=E2/D2"
    ws.cell(row=2, column=6).font = formula_font
    ws.cell(row=2, column=6).fill = formula_fill
    ws.cell(row=2, column=6).alignment = Alignment(horizontal='right', vertical='center')

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # 设置标题行高
    ws.row_dimensions[1].height = 25

    # 保存文件
    wb.save('template_with_styles.xlsx')
    wb.close()

    print("✓ 创建带样式的模板文件: template_with_styles.xlsx")
    print("  标题行样式: 蓝色背景 + 白色粗体字")
    print("  数据列样式: 浅蓝色背景 + 右对齐")
    print("  公式列样式: 浅红色背景 + 红色粗体字")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试样式保留功能")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [
                {'source': 'Sales', 'target': 'Sales'},
                {'source': 'Cost', 'target': 'Cost'},
                {'source': 'Tax', 'target': 'Tax'}
            ],
            'alias': 'external_data'
        }
    ]

    # 公式列
    formula_columns = ['SubTotal', 'Profit', 'Ratio']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_with_styles.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_styles_output.xlsx',
            use_external_refs=True
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_styles_output.xlsx")

        # 验证输出文件中的样式
        print("\n正在验证输出文件中的样式...")
        wb = openpyxl.load_workbook('test_styles_output.xlsx', data_only=False)
        ws = wb.active

        # 验证标题行样式
        header_cell = ws.cell(row=1, column=1)
        print(f"\n标题行样式验证:")
        print(f"  字体名称: {header_cell.font.name}")
        print(f"  字体大小: {header_cell.font.size}")
        print(f"  字体粗体: {header_cell.font.bold}")
        print(f"  字体颜色: {header_cell.font.color}")
        print(f"  填充颜色: {header_cell.fill}")

        # 验证数据行样式
        data_cell = ws.cell(row=2, column=1)
        print(f"\n数据列样式验证:")
        print(f"  字体名称: {data_cell.font.name}")
        print(f"  字体大小: {data_cell.font.size}")
        print(f"  对齐方式: {data_cell.alignment.horizontal}")

        # 验证公式列样式
        formula_cell = ws.cell(row=2, column=4)
        print(f"\n公式列样式验证:")
        print(f"  字体颜色: {formula_cell.font.color}")
        print(f"  字体粗体: {formula_cell.font.bold}")
        print(f"  填充颜色: {formula_cell.fill}")

        # 验证列宽
        print(f"\n列宽验证:")
        print(f"  A列宽度: {ws.column_dimensions['A'].width}")
        print(f"  D列宽度: {ws.column_dimensions['D'].width}")

        # 验证行高
        print(f"\n行高验证:")
        print(f"  标题行高度: {ws.row_dimensions[1].height}")

        wb.close()

        print("\n✓ 样式验证完成！请打开 test_styles_output.xlsx 查看效果。")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_styles()

    # 运行测试
    run_test()
