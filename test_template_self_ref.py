"""
测试模板自引用功能
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Sales': [100, 200, 300, 400, 500],
        'Cost': [50, 100, 150, 200, 250],
        'Tax': [10, 20, 30, 40, 50]
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含自引用公式的模板文件
def create_template_file():
    """创建模板文件 template_with_self_ref.xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'  # 设置模板 sheet 名称

    # 第一行：列名
    headers = ['Sales', 'Cost', 'Tax', 'SubTotal', 'Profit', 'Ratio']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # 第二行：公式模板
    # Sales, Cost, Tax 列：不设置公式（会从数据源填充）
    # SubTotal: =Sales-Cost-Tax（自引用，引用同一 sheet 的列）
    ws.cell(row=2, column=4).value = '=A2-B2-C2'

    # Profit: =SubTotal*0.2（自引用，引用 SubTotal 列）
    ws.cell(row=2, column=5).value = '=D2*0.2'

    # Ratio: =Profit/SubTotal（自引用，引用 Profit 和 SubTotal）
    ws.cell(row=2, column=6).value = '=E2/D2'

    # 保存文件
    wb.save('template_with_self_ref.xlsx')
    wb.close()

    print("✓ 创建模板文件: template_with_self_ref.xlsx")
    print("  - Sheet 名称: TemplateSheet")
    print("  - 包含自引用公式:")
    print("    - SubTotal = Sales - Cost - Tax")
    print("    - Profit = SubTotal * 0.2")
    print("    - Ratio = Profit / SubTotal")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试模板自引用功能")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [
                {'source': 'Sales', 'target': 'Sales'},
                {'source': 'Cost', 'target': 'Cost'},
                {'source': 'Tax', 'target': 'Tax'}
            ],
            'alias': 'external_data'
        }
    ]

    # 公式列（这些列的公式会被应用到输出文件）
    formula_columns = ['SubTotal', 'Profit', 'Ratio']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_with_self_ref.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_self_ref_output.xlsx',
            use_external_refs=True  # 使用外部引用模式
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_self_ref_output.xlsx")
        print("\n预期结果:")
        print("  - SubTotal 列应该包含公式: '结果'!A2-'结果'!B2-'结果'!C2")
        print("  - Profit 列应该包含公式: '结果'!D2*0.2")
        print("  - Ratio 列应该包含公式: '结果'!E2/'结果'!D2")
        print("\n注意: 公式中应该使用本地引用 '结果'!A2，而不是 [xxx.xlsx]结果!A2")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 清理旧文件
    import os
    for f in ['external_data.xlsx', 'template_with_self_ref.xlsx', 'test_self_ref_output.xlsx']:
        if os.path.exists(f):
            os.remove(f)
            print(f"清理旧文件: {f}")

    # 创建测试文件
    create_external_data_file()
    create_template_file()

    # 运行测试
    run_test()
