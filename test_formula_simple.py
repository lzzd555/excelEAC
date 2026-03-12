#!/usr/bin/env python3
"""
测试表融合功能对公式的处理
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def create_formula_excel():
    """创建包含公式的Excel文件"""
    wb = Workbook()

    # 创建表A - 包含公式列
    ws_a = wb.active
    ws_a.title = "表A"
    ws_a.append(["ID", "姓名", "基本工资", "奖金", "总收入"])
    ws_a.append(["001", "张三", 5000, "=B2*0.2", "=B2+C2"])  # 总收入=基本工资+奖金
    ws_a.append(["002", "李四", 6000, "=B3*0.2", "=B3+C3"])
    ws_a.append(["003", "王五", 7000, "=B4*0.2", "=B4+C4"])

    # 创建表B - 包含公式列
    ws_b = wb.create_sheet("表B", 1)
    ws_b.append(["员工编号", "部门", "年终奖", "总薪酬"])
    ws_b.append(["001", "销售部", "=B2*2", "=D2+E2"])
    ws_b.append(["002", "市场部", "=B3*1.5", "=D3+E3"])
    ws_b.append(["004", "技术部", 10000, 20000])

    wb.save("test_formulas.xlsx")
    print("公式测试文件已创建：test_formulas.xlsx")

def read_with_pandas():
    """使用pandas读取文件查看数据"""
    try:
        print("=== 使用pandas读取Excel文件 ===")

        # 读取表A
        df_a = pd.read_excel("test_formulas.xlsx", sheet_name="表A")
        print("表A数据：")
        print(df_a)

        # 读取表B
        df_b = pd.read_excel("test_formulas.xlsx", sheet_name="表B")
        print("\n表B数据：")
        print(df_b)

        return df_a, df_b

    except Exception as e:
        print(f"读取失败: {e}")
        return None, None

def read_with_openpyxl():
    """使用openpyxl读取文件查看公式和值"""
    try:
        from openpyxl import load_workbook

        print("\n=== 使用openpyxl读取Excel文件（查看公式和值） ===")

        # 读取表A
        wb_a = load_workbook("test_formulas.xlsx", data_only=False)
        ws_a = wb_a["表A"]
        print("表A（公式）：")
        for row in ws_a.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                if cell.data_type == 'f':  # 公式
                    row_data.append(f"'{cell.value}")
                else:
                    row_data.append(cell.value)
            print(row_data)

        # 读取表B
        wb_b = load_workbook("test_formulas.xlsx", data_only=False)
        ws_b = wb_b["表B"]
        print("\n表B（公式）：")
        for row in ws_b.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                if cell.data_type == 'f':  # 公式
                    row_data.append(f"'{cell.value}")
                else:
                    row_data.append(cell.value)
            print(row_data)

        # 读取计算后的值
        print("\n=== 读取计算后的值 ===")
        wb_values = load_workbook("test_formulas.xlsx", data_only=True)

        # 表A的值
        print("表A（值）：")
        for row in wb_values["表A"].iter_rows(values_only=True):
            print(row)

        # 表B的值
        print("\n表B（值）：")
        for row in wb_values["表B"].iter_rows(values_only=True):
            print(row)

    except Exception as e:
        print(f"openpyxl读取失败: {e}")

if __name__ == "__main__":
    # 创建测试文件
    create_formula_excel()

    # 读取文件
    df_a, df_b = read_with_pandas()
    read_with_openpyxl()

    # 如果pandas能读到数据，尝试合并
    if df_a is not None and df_b is not None:
        print("\n=== 尝试使用表融合功能 ===")
        try:
            from modules.merge import merge_excel_tables

            result = merge_excel_tables(
                table_a_file="test_formulas.xlsx",
                table_a_sheet="表A",
                table_b_file="test_formulas.xlsx",
                table_b_sheet="表B",
                match_columns={'ID': '员工编号'},
                table_a_extra_columns=['姓名', '基本工资', '奖金', '总收入'],
                table_b_extra_columns=['部门', '年终奖', '总薪酬'],
                output_file='formula_merge_result.xlsx',
                string_columns=['ID', '员工编号']
            )

            print("\n合并结果：")
            print(result)

            # 检查输出文件中的公式
            print("\n=== 检查输出文件中的内容 ===")
            from openpyxl import load_workbook
            wb_out = load_workbook("formula_merge_result.xlsx", data_only=False)
            ws_out = wb_out["合并结果"]

            print("输出表（公式）：")
            for row in ws_out.iter_rows(values_only=False):
                row_data = []
                for cell in row:
                    if cell.data_type == 'f':  # 公式
                        row_data.append(f"'{cell.value}")
                    else:
                        row_data.append(cell.value)
                print(row_data)

            # 读取值
            print("\n输出表（值）：")
            wb_values_out = load_workbook("formula_merge_result.xlsx", data_only=True)
            for row in wb_values_out["合并结果"].iter_rows(values_only=True):
                print(row)

        except Exception as e:
            print(f"合并失败: {e}")