"""
测试多数据源的外部引用映射
"""
import pandas as pd
import openpyxl
import os

# 清理旧文件
for f in ['data1.xlsx', 'data2.xlsx', 'data3.xlsx', 'template_multi_sources.xlsx', 'test_multi_sources_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建多个数据源文件
def create_data_files():
    """创建三个数据源文件"""
    # 数据源 1
    data1 = pd.DataFrame({'Value1': [10, 20, 30]})
    with pd.ExcelWriter('data1.xlsx', engine='openpyxl') as writer:
        data1.to_excel(writer, sheet_name='Sheet1', index=False)
    print("✓ 创建 data1.xlsx (Sheet1)")

    # 数据源 2
    data2 = pd.DataFrame({'Value2': [100, 200, 300]})
    with pd.ExcelWriter('data2.xlsx', engine='openpyxl') as writer:
        data2.to_excel(writer, sheet_name='Sheet2', index=False)
    print("✓ 创建 data2.xlsx (Sheet2)")

    # 数据源 3
    data3 = pd.DataFrame({'Value3': [1000, 2000, 3000]})
    with pd.ExcelWriter('data3.xlsx', engine='openpyxl') as writer:
        data3.to_excel(writer, sheet_name='Sheet3', index=False)
    print("✓ 创建 data3.xlsx (Sheet3)")


# 创建引用第二个数据源的模板
def create_template():
    """创建引用第二个数据源的模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'

    # 第一行：列名
    headers = ['Value1', 'Value2', 'Value3', 'ExternalRef', 'Total']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # 第二行：公式
    # ExternalRef: 引用第二个数据源 (应该是 data2.xlsx 的 Sheet2!A1)
    # 使用 sheet 名称引用
    ws.cell(row=2, column=4).value = "=Sheet2!A1"

    # Total: =Value1 + ExternalRef
    ws.cell(row=2, column=5).value = "=A2+D2"

    # 保存文件
    wb.save('template_multi_sources.xlsx')
    wb.close()

    print("✓ 创建模板文件: template_multi_sources.xlsx")
    print("  ExternalRef 公式: =[2]Sheet2!A1 (引用第二个数据源)")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始测试：多数据源外部引用映射")
    print("="*70 + "\n")

    # 数据源配置 - 注意顺序
    data_sources = [
        {
            'file_path': 'data1.xlsx',
            'sheet_name': 'Sheet1',
            'column_mappings': [{'source': 'Value1', 'target': 'Value1'}],
            'alias': 'data1'
        },
        {
            'file_path': 'data2.xlsx',
            'sheet_name': 'Sheet2',
            'column_mappings': [{'source': 'Value2', 'target': 'Value2'}],
            'alias': 'data2'
        },
        {
            'file_path': 'data3.xlsx',
            'sheet_name': 'Sheet3',
            'column_mappings': [{'source': 'Value3', 'target': 'Value3'}],
            'alias': 'data3'
        }
    ]

    # 公式列
    formula_columns = ['ExternalRef', 'Total']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_multi_sources.xlsx',
            template_sheet='TemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_multi_sources_output.xlsx',
            use_external_refs=True
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_multi_sources_output.xlsx")

        # 验证输出文件中的公式
        print("\n正在验证输出文件中的公式...")
        wb = openpyxl.load_workbook('test_multi_sources_output.xlsx', data_only=False)
        ws = wb.active

        # 读取第二行的公式
        external_ref_formula = ws.cell(row=2, column=4).value
        total_formula = ws.cell(row=2, column=5).value

        print(f"\n公式验证结果:")
        print(f"  ExternalRef: {external_ref_formula}")
        print(f"  Total: {total_formula}")

        wb.close()

        # 检查引用是否正确
        print("\n结果分析:")
        print(f"  模板中的公式: =[2]Sheet2!A1")
        print(f"  期望输出公式: 应该引用 data2.xlsx 的 Sheet2")
        print(f"  实际输出公式: {external_ref_formula}")

        if 'data2.xlsx' in str(external_ref_formula) and 'Sheet2' in str(external_ref_formula):
            print("\n✓ 正确：公式引用了正确的数据源 (data2.xlsx/Sheet2)")
        else:
            print("\n✗ 错误：公式引用了错误的数据源")
            print(f"  期望: [data2.xlsx]Sheet2")
            print(f"  实际: {external_ref_formula}")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 创建测试文件
    create_data_files()
    create_template()

    # 运行测试
    run_test()
