#!/usr/bin/env python3
"""
验证修复：确保字符串格式被正确保持
"""

import pandas as pd
from excel_validator import process_excel_with_validation
import openpyxl

def verify_fix():
    print("=== 验证修复：字符串格式保持 ===\n")

    # 创建测试数据
    data = {
        '订单号': ['001', '002', '003', '004'],
        '产品代码': ['001', '002', '001', '003'],
        '部门': ['A', 'A', 'B', 'B'],
        '计划值': [100, 200, 150, 120],
        '实际值': [100, 210, 150, 120]  # 第2行有异常
    }

    df = pd.DataFrame(data)
    test_file = 'verify_test.xlsx'
    df.to_excel(test_file, index=False)

    print("1. 原始数据：")
    print(df)

    # 运行验证
    result = process_excel_with_validation(
        input_file=test_file,
        sheet_name='Sheet1',
        group_columns=['部门'],
        compare_columns=['计划值', '实际值'],
        output_columns=['部门', '验证状态', '总行数'],
        output_file='verify_output.xlsx',
        string_columns=['订单号', '产品代码']
    )

    print("\n2. 分组结果：")
    print(result)

    # 直接读取Excel文件
    print("\n3. 检查Excel文件中的异常详情...")
    wb = openpyxl.load_workbook('verify_output.xlsx')
    detail_sheet = wb['异常详情']

    # 打印所有内容
    print("\n异常详情完整内容：")
    for row in detail_sheet.iter_rows(values_only=True):
        print(row)

    # 特别检查订单号
    print("\n订单号详细检查：")
    # 找到订单号的列索引
    order_id_col_idx = None
    for i, cell in enumerate(detail_sheet[1]):
        if cell.value == '订单号':
            order_id_col_idx = i
            break

    if order_id_col_idx is not None:
        print(f"订单号在列 {order_id_col_idx + 1}")
        for i, row in enumerate(detail_sheet.iter_rows(min_row=2, values_only=True)):
            order_id = row[order_id_col_idx]
            print(f"  行 {i+1}: 订单号='{order_id}' (类型: {type(order_id)})")
            if str(order_id) == '002':
                print("    ✅ 格式正确！")
            else:
                print("    ❌ 格式错误！")
    else:
        print("❌ 没有找到订单号列")

    # 创建一个简单的pandas测试
    print("\n4. pandas string类型测试：")
    test_df = pd.DataFrame({'订单号': ['001', '002']})
    test_df['订单号'] = test_df['订单号'].astype('string')
    print("使用string类型：", test_df['订单号'].tolist())
    print("类型：", test_df['订单号'].dtype)

    test_df['订单号'] = test_df['订单号'].astype(str)
    print("使用str类型：", test_df['订单号'].tolist())
    print("类型：", test_df['订单号'].dtype)

if __name__ == "__main__":
    verify_fix()