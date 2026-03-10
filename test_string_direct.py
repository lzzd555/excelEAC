#!/usr/bin/env python3
"""
直接检查Excel文件中的字符串格式
"""

import pandas as pd
from excel_validator import process_excel_with_validation
import openpyxl

def test_string_direct():
    print("=== 直接检查Excel中的字符串格式 ===\n")

    # 创建测试数据
    data = {
        '订单号': ['001', '002', '003', '004', '005'],
        '产品代码': ['A01', 'A02', 'A01', 'A03', 'A02'],
        '部门': ['销售部', '销售部', '市场部', '市场部', '销售部'],
        '计划金额': [1000, 2000, 1500, 1800, 2200],
        '实际金额': [1000, 2100, 1500, 1800, 2150]
    }

    df = pd.DataFrame(data)
    test_file = 'string_test.xlsx'
    df.to_excel(test_file, index=False)

    print("原始数据订单号:", df['订单号'].tolist())

    # 运行验证
    result = process_excel_with_validation(
        input_file=test_file,
        sheet_name='Sheet1',
        group_columns=['部门'],
        compare_columns=['计划金额', '实际金额'],
        output_columns=['部门', '验证状态'],
        output_file='string_output.xlsx',
        string_columns=['订单号', '产品代码'],
        abnormal_detail_columns=['订单号', '产品代码', '计划金额', '实际金额']
    )

    print("分组结果:")
    print(result)

    # 直接从Excel文件读取异常详情
    wb = openpyxl.load_workbook('string_output.xlsx')
    detail_sheet = wb['异常详情']

    print("\n异常详情（直接从Excel读取）:")
    for row in detail_sheet.iter_rows(values_only=True):
        print(row)

    # 找到订单号列
    order_col_idx = None
    for i, cell in enumerate(detail_sheet[1]):
        if cell.value == '订单号':
            order_col_idx = i
            break

    if order_col_idx is not None:
        print(f"\n订单号在列 {order_col_idx + 1}")
        order_ids = []
        for row in detail_sheet.iter_rows(min_row=2, values_only=True):
            order_ids.append(row[order_col_idx])

        print("Excel中的订单号:", order_ids)

        # 检查前导零
        has_leading_zeros = any(str(id).startswith('00') for id in order_ids)
        if has_leading_zeros:
            print("✅ 订单号前导零保持正常")
        else:
            print("❌ 订单号前导零丢失")

    # 使用pandas读取并检查
    detail_df = pd.read_excel('string_output.xlsx', sheet_name='异常详情')
    print("\n使用pandas读取的订单号:", detail_df['订单号'].tolist())
    print("pandas数据类型:", detail_df['订单号'].dtype)

if __name__ == "__main__":
    test_string_direct()