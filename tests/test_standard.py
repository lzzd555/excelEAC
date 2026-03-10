#!/usr/bin/env python3
"""
标准测试用例：验证excel_validator.py的功能
基于用户修改后的test_abnormal_detail.py
"""

import pandas as pd
from excel_validator import process_excel_with_validation
import openpyxl
import os

def standard_test():
    """标准测试用例"""
    print("=== 标准测试用例：验证excel_validator.py功能 ===\n")

    # 测试数据（使用用户提供的测试数据）
    data = {
        '订单号': ['001', '002', '003', '004'],
        '产品代码': ['001', '002', '001', '003'],
        '部门': ['A', 'A', 'B', 'B'],
        '计划值': [100, 200, 150, 120],
        '实际值': [100, 210, 150, 120]  # ORD002 有异常
    }

    df = pd.DataFrame(data)
    test_file = 'standard_test.xlsx'
    df.to_excel(test_file, index=False)

    print("1. 测试数据：")
    print(df)
    print("\n数据类型检查：")
    for col in df.columns:
        print(f"  {col}: {df[col].dtype}")

    print("\n2. 运行验证...")
    result = process_excel_with_validation(
        input_file=test_file,
        sheet_name='Sheet1',
        group_columns=['部门'],
        compare_columns=['计划值', '实际值'],
        output_columns=['部门', '验证状态', '总行数'],
        output_file='standard_test_output.xlsx',
        string_columns=['订单号', '产品代码']
    )

    print("\n3. 分组结果：")
    print(result)
    print("\n分组结果类型检查：")
    for col in result.columns:
        print(f"  {col}: {result[col].dtype}")

    print("\n4. 检查异常详情...")
    output_file = 'standard_test_output.xlsx'
    try:
        wb = openpyxl.load_workbook(output_file)

        # 检查所有sheet
        print(f"Excel文件包含的sheet: {wb.sheetnames}")

        # 检查异常详情
        if '异常详情' in wb.sheetnames:
            detail_sheet = wb['异常详情']
            print("\n异常详情内容：")

            # 获取列名
            header = []
            for cell in detail_sheet[1]:  # 第一行是列名
                header.append(str(cell.value))
            print(f"列名: {header}")

            # 首先找到订单号的列索引
            order_col_idx = None
            for i, col_name in enumerate(detail_sheet[1]):
                if col_name == '订单号':
                    order_col_idx = i
                    break

            # 获取数据行
            data_rows = []
            order_id_values = []
            for row in detail_sheet.iter_rows(min_row=2, values_only=True):
                data_rows.append(row)
                if order_col_idx is not None and row[order_col_idx]:
                    order_id_values.append(row[order_col_idx])

            print(f"\n数据行数: {len(data_rows)}")
            print("前5行数据：")
            for i, row in enumerate(data_rows[:5]):
                print(f"  行{i+1}: {row}")

            # 重点检查订单号格式
            print(f"\n订单号检查：")
            print(f"  订单号值: {order_id_values}")
            print(f"  订单号类型: {[type(id) for id in order_id_values]}")

            # 检查是否保持了正确的格式
            if '002' in order_id_values:
                print("  ✅ 订单号格式正确！保持了'002'格式")
            else:
                print("  ❌ 订单号格式有误！")

            # 检查订单号在Excel中的实际格式
            print("\nExcel中订单号的实际格式检查：")
            order_id_col_idx = None
            for i, col_name in enumerate(detail_sheet[1]):
                if col_name == '订单号':
                    order_id_col_idx = i
                    break

            if order_id_col_idx is not None:
                print(f"订单号列索引: {order_id_col_idx}")
                for i, row in enumerate(detail_sheet.iter_rows(min_row=2, values_only=True)):
                    if row:  # 非空行
                        order_id = row[order_id_col_idx]
                        print(f"  第{i+1}行订单号: '{order_id}' (类型: {type(order_id)})")
                        if str(order_id) in ['001', '002', '003', '004']:
                            print(f"    ✅ 格式正确: '{order_id}'")
                        else:
                            print(f"    ❌ 格式可能有问题: '{order_id}'")
        else:
            print("❌ 没有异常详情sheet")

        # 检查验证结果sheet
        if '验证结果' in wb.sheetnames:
            result_sheet = wb['验证结果']
            print("\n验证结果内容：")
            for row in result_sheet.iter_rows(values_only=True):
                print(row)

    except Exception as e:
        print(f"读取Excel文件失败: {e}")

    print("\n5. 测试总结：")
    print("关键点验证：")
    print("- 订单号是否保持'001'格式（而不是'1'）")
    print("- 产品代码是否保持'001'格式")
    print("- 异常详情是否包含正确的数据")
    print("- 分组结果是否正确")

    # 不自动删除文件，便于用户检查
    # print(f"\n清理临时文件...")
    # os.remove(test_file)
    # os.remove(output_file)
    print(f"\n临时文件已保留：")
    print(f"  输入文件: {test_file}")
    print(f"  输出文件: {output_file}")
    print("请手动删除这些文件。")

if __name__ == "__main__":
    standard_test()