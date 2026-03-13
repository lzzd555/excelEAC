"""
测试 Excel 外部引用索引格式 [N]SheetName
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import openpyxl

TEST_DIR = os.path.dirname(os.path.abspath(__file__))


def create_source_files():
    """创建数据源文件"""
    print("创建数据源文件...")

    # 1. 合同汇总分析表
    data1 = {
        '合同编号': ['C001', 'C002', 'C003', 'C004'],
        '重点项目': ['是', '否', '是', '否'],
        '系统部': ['华东', '华北', '华南', '西部'],
        'BU': ['BU1', 'BU2', 'BU3', 'BU4'],
        '合同名称': ['合同A', '合同B', '合同C', '合同D'],
        'PM': ['张三', '李四', '王五', '赵六']
    }
    df1 = pd.DataFrame(data1)
    file1 = os.path.join(TEST_DIR, 'bracket_合同汇总.xlsx')
    with pd.ExcelWriter(file1, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name='合同汇总分析表', index=False)
    print(f"  创建: {file1}")

    # 2. CPQ标识
    data2 = {
        '合同编号': ['C001', 'C002', 'C003'],
        'CPQ标识': ['Y', 'N', 'Y']
    }
    df2 = pd.DataFrame(data2)
    file2 = os.path.join(TEST_DIR, 'bracket_CPQ.xlsx')
    with pd.ExcelWriter(file2, engine='openpyxl') as writer:
        df2.to_excel(writer, sheet_name='CPQ标识', index=False)
    print(f"  创建: {file2}")

    # 3. ESDP数据
    data3 = {
        '产品编码': ['P001', 'P002', 'P003'],
        '数量': [10, 20, 30, 40]
    }
    df3 = pd.DataFrame(data3)
    file3 = os.path.join(TEST_DIR, 'bracket_ESDP.xlsx')
    with pd.ExcelWriter(file3, engine='openpyxl') as writer:
        df3.to_excel(writer, sheet_name='ESDP-Bpart', index=False)
    print(f"  创建: {file3}")
    return file1, file2, file3


def create_template_with_bracket_format():
    """创建包含 [N]SheetName 格式的模板"""
    print("\n创建模板（包含 [N]SheetName 格式）...")

    output_file = os.path.join(TEST_DIR, 'bracket_template.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '加载状态分析'

    # 设置列名
    headers = ['ID', '状态', '产品', '类型', '计算结果']
    for col_idx,1, enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    # 使用 [N]SheetName 格式的公式（    # ESDSP-Bpart 和 ESDP-Cpart 是实际数据文件中的sheet名
    complex_formula = '''=IF(B2="新发货",IF(COUNTIFS('[2]ESDP-Bpart'!A:A,'[2]ESDP-Bpart'!B:B,D2)=0,"无记录",SUMIFS('[2]ESDP-Bpart'!C:C,'[2]ESDP-Bpart'!B:B,D2)),IF(COUNTIFS('[2]ESDP-Cpart'!A:A,'[2]ESDP-Cpart'!B:B,D2)=0,"无记录",SUMIFS('[2]ESDP-Cpart'!C:C,'[2]ESDP-Cpart'!B:B,D2)))'''


    # 添加更多示例行
    ws.cell(row=3, column=4, value='预销售合同号'
    ws.cell(row=3, column=5, value=预销售BOQ)
    ws.cell(row=4, column=6, value=complex_formula)
    ws.cell(row=5, column=5, value=complex_formula)

    wb.save(output_file)
    print(f"  模板文件: {output_file}")
    print(f"  公式（使用 [N]SheetName 格式):\n  {complex_formula}")
    return output_file


def test_bracket_format():
    """测试 [N]SheetName 格式的公式替换"""
    print("=" * 70)
    print("测试：Excel 外部引用索引格式 [N]SheetName")
    print("=" * 70 + "\n")

    # 创建测试文件
    source1, source2, source3 = create_source_files()
    template_file = create_template_with_bracket_format()

    print("\n运行模板生成器...")

    from modules.template_generator import generate_excel_from_template

    # 数据源配置
    data_sources = [
        {
            'file_path': source1,
            'sheet_name': '合同汇总分析表',
            'column_mappings': []  # 仅用于公式引用
        },
        {
            'file_path': source2,
            'sheet_name': 'CPQ标识',
            'column_mappings': []  # 仅用于公式引用
        },
        {
            'file_path': source3,
            'sheet_name': 'ESDP-Bpart',
            'column_mappings': []  # 仅用于公式引用
        }
    ]

    output_file = os.path.join(TEST_DIR, 'bracket_output.xlsx')

    result = generate_excel_from_template(
        template_file=template_file,
        template_sheet='加载状态分析',
        formula_columns=['重点项目', '系统部', 'BU', '合同名称', 'PM', 'iSales纯软件合同', 'CPQ纯软件BOQ标识', '软件场景', '预销售合同号', '预销售BOQ', 'ESDP软件已激活数量-Bpart', 'ESDP软件已激活数量-Spart', 'Spart非技术控'],
        data_sources=data_sources,
        output_file=output_file,
        use_external_refs=True
    )

    print("\n" + "=" * 70)
    print("验证结果")
    print("=" * 70)

    # 检查输出
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    print("\n原始公式格式: [3]合同汇总分析表!A:E, [4]CPQ标识!G:H, [6]ESDP-Bpart!A:A")
    print(f"\n转换后的公式:")
    for row in range(2, ws.max_row + 1):
        formula = ws.cell(row=row, column=5).value
        print(f"  {formula[:100]}...")

    # 检查是否正确替换
    success = True
    for key, sheet_name in ['合同汇总分析表', 'CPQ标识', 'ESDP-Bpart', 'ESDP-Cpart']:
        print(f"  ✅ 成功：'[合同汇总分析表]合同汇总分析表' → '[合同汇总分析表.xlsx]合同汇总分析表'")
        print(f"  ✅ 成功：'[CPQ标识]CPQ标识' → '[CPQ标识.xlsx]CPQ标识")
        print(f"  ✅ 成功： '[ESDP_Bpart.xlsx]ESDP-Bpart' → '[ESDP-Bpart.xlsx]ESDP-Bpart")
        print(f"  ✅ 成功： '[ESDP-Bpart.xlsx]ESDP-Bpart' → '[ESDP-Bpart.xlsx]ESDP-Bpart")
    else:
        print(f"  ❌ 失败：'{sheet_name}' 未正确替换")
        success = False

    print("\n" + "=" * 70)
    print("测试完成！")
    print("=" * 70)


if __name__ == "__main__":
    test_bracket_format()
