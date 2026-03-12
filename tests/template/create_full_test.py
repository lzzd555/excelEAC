"""
创建完整的模板生成器测试场景
包含：公式模板、销售数据、成本数据
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

TEST_DIR = os.path.dirname(os.path.abspath(__file__))


def create_sales_data():
    """创建销售数据文件"""
    print("创建销售数据...")

    data = {
        '日期': ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04', '2024-01-05', '2024-01-06'],
        '产品': ['A001', 'A002', 'A001', 'A003', 'A002', 'A001'],
        '销售额': [10000, 15000, 12000, 8000, 20000, 18000],
        '数量': [100, 150, 120, 80, 200, 180],
        '销售员': ['张三', '李四', '张三', '王五', '李四', '张三']
    }
    df = pd.DataFrame(data)
    output_file = os.path.join(TEST_DIR, 'test_sales.xlsx')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='销售数据', index=False)

        # 格式化
        ws = writer.sheets['销售数据']
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12

    print(f"  文件: {output_file}")
    print(f"  数据:\n{df}\n")
    return output_file


def create_costs_data():
    """创建成本数据文件"""
    print("创建成本数据...")

    data = {
        '日期': ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04', '2024-01-05'],
        '产品': ['A001', 'A002', 'A001', 'A003', 'A002'],
        '成本额': [5000, 8000, 6000, 4000, 10000],
        '运费': [200, 300, 240, 160, 400]
    }
    df = pd.DataFrame(data)
    output_file = os.path.join(TEST_DIR, 'test_costs.xlsx')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='成本数据', index=False)

        # 格式化
        ws = writer.sheets['成本数据']
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12

    print(f"  文件: {output_file}")
    print(f"  数据:\n{df}\n")
    return output_file


def create_template():
    """创建带有公式的模板文件"""
    print("创建模板文件...")

    output_file = os.path.join(TEST_DIR, 'test_template.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '报表模板'

    # 设置标题行样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # 列名
    headers = ['日期', '产品', '销售额', '成本额', '毛利', '毛利率', '运费', '净利']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 第二行设置公式示例（会被数据覆盖，仅用于记录公式模板）
    # 销售额从sheet0获取，成本额从sheet1获取
    ws.cell(row=2, column=1, value='示例日期')
    ws.cell(row=2, column=2, value='示例产品')
    ws.cell(row=2, column=3, value='=sheet0!C2')      # 销售额 - 从销售表的C列
    ws.cell(row=2, column=4, value='=sheet1!C2')      # 成本额 - 从成本表的C列
    ws.cell(row=2, column=5, value='=C2-D2')          # 毛利 = 销售额 - 成本额
    ws.cell(row=2, column=6, value='=E2/C2')          # 毛利率 = 毛利/销售额
    ws.cell(row=2, column=7, value='=sheet1!D2')      # 运费 - 从成本表的D列
    ws.cell(row=2, column=8, value='=E2-G2')          # 净利 = 毛利 - 运费

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12

    wb.save(output_file)

    print(f"  文件: {output_file}")
    print(f"  公式列: 销售额, 成本额, 毛利, 毛利率, 运费, 净利")
    print(f"  公式示例:")
    print(f"    销售额: =sheet0!C2 (引用销售数据)")
    print(f"    成本额: =sheet1!C2 (引用成本数据)")
    print(f"    毛利: =C2-D2 (销售额-成本额)")
    print(f"    毛利率: =E2/C2 (毛利/销售额)")
    print(f"    运费: =sheet1!D2 (引用成本数据)")
    print(f"    净利: =E2-G2 (毛利-运费)\n")

    return output_file


def main():
    print("=" * 60)
    print("创建模板生成器测试场景")
    print("=" * 60 + "\n")

    sales_file = create_sales_data()
    costs_file = create_costs_data()
    template_file = create_template()

    print("=" * 60)
    print("测试文件创建完成！")
    print("=" * 60)
    print(f"\n模板文件: {template_file}")
    print(f"销售数据: {sales_file}")
    print(f"成本数据: {costs_file}")

    print("\n" + "=" * 60)
    print("运行命令示例:")
    print("=" * 60)
    cmd = f'''python main.py template \\
    -t tests/template/test_template.xlsx \\
    -ts 报表模板 \\
    -f "销售额,成本额,毛利,毛利率,运费,净利" \\
    -d tests/template/test_sales.xlsx 销售数据 "日期:日期,产品:产品,销售额:销售额" sheet0 \\
    -d tests/template/test_costs.xlsx 成本数据 "日期:日期,产品:产品,成本额:成本额,运费:运费" sheet1 \\
    -o tests/template/test_output.xlsx \\
    --string-columns "日期,产品"
'''
    print(cmd)

    return template_file, sales_file, costs_file


if __name__ == "__main__":
    main()
