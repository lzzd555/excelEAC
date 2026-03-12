"""
创建模板生成器测试数据
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os

# 测试目录
TEST_DIR = os.path.dirname(os.path.abspath(__file__))


def create_sales_data():
    """创建销售数据文件"""
    data = {
        'Date': ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04', '2024-01-05'],
        'SalesAmt': [1000, 1500, 2000, 1200, 1800],
        'Region': ['East', 'West', 'East', 'North', 'South']
    }
    df = pd.DataFrame(data)
    output_file = os.path.join(TEST_DIR, 'sales.xlsx')
    df.to_excel(output_file, sheet_name='Data', index=False)
    print(f"创建销售数据: {output_file}")
    return output_file


def create_costs_data():
    """创建成本数据文件"""
    data = {
        'Date': ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04'],
        'CostAmt': [500, 600, 800, 550],
        'Category': ['A', 'B', 'A', 'C']
    }
    df = pd.DataFrame(data)
    output_file = os.path.join(TEST_DIR, 'costs.xlsx')
    df.to_excel(output_file, sheet_name='Data', index=False)
    print(f"创建成本数据: {output_file}")
    return output_file


def create_template():
    """创建模板文件（包含公式）"""
    output_file = os.path.join(TEST_DIR, 'template.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 设置列名
    headers = ['Date', 'Sales', 'Cost', 'Total', 'Profit']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 设置第二行为公式示例
    # Date列
    ws.cell(row=2, column=1, value='2024-01-01')
    # Sales列
    ws.cell(row=2, column=2, value=1000)
    # Cost列
    ws.cell(row=2, column=3, value=500)
    # Total公式: =Sales+Cost (非sheet引用公式)
    ws.cell(row=2, column=4, value='=B2+C2')
    # Profit公式: =sheet0!B2-sheet1!B2 (sheet引用公式)
    # sheet0对应sales数据，sheet1对应costs数据
    # B列是第2列，所以是B2
    ws.cell(row=2, column=5, value='=sheet0!B2-sheet1!B2')

    wb.save(output_file)
    print(f"创建模板文件: {output_file}")
    return output_file


def create_template_with_external_ref():
    """
    创建包含外部文件引用的模板文件
    公式格式: =[filename.xlsx]sheetname!A1
    """
    output_file = os.path.join(TEST_DIR, 'template_external.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 设置列名
    headers = ['Date', 'Sales', 'Cost', 'Profit']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 设置公式示例（使用sheet别名）
    # sheet0 -> sales.xlsx
    # sheet1 -> costs.xlsx
    # Sales从sheet0取值，Cost从sheet1取值
    ws.cell(row=2, column=1, value='2024-01-01')
    ws.cell(row=2, column=2, value='=sheet0!B2')  # Sales
    ws.cell(row=2, column=3, value='=sheet1!B2')  # Cost
    ws.cell(row=2, column=4, value='=B2-C2')      # Profit = Sales - Cost

    wb.save(output_file)
    print(f"创建外部引用模板文件: {output_file}")
    return output_file


def main():
    print("=== 创建模板生成器测试数据 ===\n")

    create_sales_data()
    create_costs_data()
    create_template()
    create_template_with_external_ref()

    print("\n所有测试数据创建完成！")


if __name__ == "__main__":
    main()
