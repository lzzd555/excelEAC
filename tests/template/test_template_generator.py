"""
测试模板生成器功能
"""

import sys
import os

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
from modules.template_generator import (
    generate_excel_from_template,
    parse_column_mappings,
    read_template_structure,
    parse_formula_references,
    replace_sheet_references
)


def test_parse_column_mappings():
    """测试列映射解析"""
    print("=== 测试列映射解析 ===")

    # 测试带冒号的格式
    result = parse_column_mappings("SalesAmt:Sales,CostAmt:Cost")
    assert len(result) == 2
    assert result[0] == {'source': 'SalesAmt', 'target': 'Sales'}
    assert result[1] == {'source': 'CostAmt', 'target': 'Cost'}
    print("   带冒号格式: 通过")

    # 测试不带冒号的格式
    result = parse_column_mappings("Date,Region")
    assert len(result) == 2
    assert result[0] == {'source': 'Date', 'target': 'Date'}
    assert result[1] == {'source': 'Region', 'target': 'Region'}
    print("   不带冒号格式: 通过")

    print()


def test_parse_formula_references():
    """测试公式引用解析"""
    print("=== 测试公式引用解析 ===")

    # 测试简单公式
    formula = "=sheet0!A1+sheet1!B2"
    refs = parse_formula_references(formula)
    assert len(refs) == 2
    assert refs[0] == ('sheet0', 'A', 1)
    assert refs[1] == ('sheet1', 'B', 2)
    print("   简单公式: 通过")

    # 测试复杂公式
    formula = "=sheet0!A1*2+sheet1!C10/100"
    refs = parse_formula_references(formula)
    assert len(refs) == 2
    print("   复杂公式: 通过")

    print()


def test_replace_sheet_references():
    """测试公式引用替换"""
    print("=== 测试公式引用替换 ===")

    alias_to_info = {
        'sheet0': {'file_path': '/path/to/sales.xlsx', 'sheet_name': '销售数据'},
        'sheet1': {'file_path': '/path/to/costs.xlsx', 'sheet_name': '成本数据'}
    }

    # 测试简单替换
    formula = "=sheet0!A1+sheet1!B2"
    result = replace_sheet_references(formula, alias_to_info, row_offset=0)
    assert "'[sales.xlsx]销售数据'!A1" in result
    assert "'[costs.xlsx]成本数据'!B2" in result
    print(f"   简单替换: {result}")

    # 测试带行偏移的替换
    formula = "=sheet0!A1"
    result = replace_sheet_references(formula, alias_to_info, row_offset=5)
    assert "'[sales.xlsx]销售数据'!A6" in result
    print(f"   带行偏移替换: {result}")

    # 测试本地单元格引用的行偏移
    formula = "=B2-C2"
    result = replace_sheet_references(formula, alias_to_info, row_offset=0)
    assert result == "=B2-C2"
    print(f"   本地引用(无偏移): {result}")

    formula = "=B2-C2"
    result = replace_sheet_references(formula, alias_to_info, row_offset=3)
    assert result == "=B5-C5"
    print(f"   本地引用(偏移3): {result}")

    # 测试混合公式
    formula = "=sheet0!B2-C2"
    result = replace_sheet_references(formula, alias_to_info, row_offset=1)
    assert "'[sales.xlsx]销售数据'!B3" in result
    assert "C3" in result
    print(f"   混合公式: {result}")

    print()


def test_full_generation():
    """测试完整的模板生成流程"""
    print("=== 测试完整模板生成 ===")

    test_dir = os.path.dirname(os.path.abspath(__file__))

    # 首先创建测试数据
    from create_test_data import create_sales_data, create_costs_data, create_template_with_external_ref

    sales_file = create_sales_data()
    costs_file = create_costs_data()
    template_file = create_template_with_external_ref()

    output_file = os.path.join(test_dir, 'output.xlsx')

    # 准备数据源配置
    data_sources = [
        {
            'file_path': sales_file,
            'sheet_name': 'Data',
            'column_mappings': [
                {'source': 'Date', 'target': 'Date'},
                {'source': 'SalesAmt', 'target': 'Sales'}
            ],
            'alias': 'sheet0'
        },
        {
            'file_path': costs_file,
            'sheet_name': 'Data',
            'column_mappings': [
                {'source': 'Date', 'target': 'Date'},
                {'source': 'CostAmt', 'target': 'Cost'}
            ],
            'alias': 'sheet1'
        }
    ]

    # 执行模板生成
    result = generate_excel_from_template(
        template_file=template_file,
        template_sheet='Sheet1',
        formula_columns=['Sales', 'Cost', 'Profit'],  # 这些列使用公式
        data_sources=data_sources,
        output_file=output_file,
        string_columns=['Date']
    )

    print(f"\n生成结果:")
    print(result)

    # 验证输出文件
    assert os.path.exists(output_file), "输出文件不存在"
    print(f"\n输出文件已创建: {output_file}")

    # 读取输出文件验证
    import openpyxl
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # 检查公式
    print("\n检查公式:")
    for row in range(2, ws.max_row + 1):
        sales_cell = ws.cell(row=row, column=2)
        cost_cell = ws.cell(row=row, column=3)
        profit_cell = ws.cell(row=row, column=4)

        print(f"  行{row}: Sales={sales_cell.value}, Cost={cost_cell.value}, Profit={profit_cell.value}")

    wb.close()

    print("\n完整测试通过！")


def test_cli():
    """测试命令行接口"""
    print("\n=== 测试命令行接口 ===")

    test_dir = os.path.dirname(os.path.abspath(__file__))

    # 构建命令行
    cmd = f"""python {os.path.dirname(os.path.dirname(test_dir))}/main.py template \\
        -t {test_dir}/template_external.xlsx \\
        -ts Sheet1 \\
        -f "Sales,Cost,Profit" \\
        -d {test_dir}/sales.xlsx Data "Date:Date,SalesAmt:Sales" sheet0 \\
        -d {test_dir}/costs.xlsx Data "Date:Date,CostAmt:Cost" sheet1 \\
        -o {test_dir}/cli_output.xlsx"""

    print(f"命令:\n{cmd}")
    print("\n请手动运行上述命令进行测试")


def test_internal_vs_external_refs():
    """测试内部引用模式和外部引用模式的区别"""
    print("\n=== 测试内部引用 vs 外部引用模式 ===")

    test_dir = os.path.dirname(os.path.abspath(__file__))
    from create_test_data import create_sales_data, create_costs_data, create_template_with_external_ref

    sales_file = create_sales_data()
    costs_file = create_costs_data()
    template_file = create_template_with_external_ref()

    data_sources = [
        {
            'file_path': sales_file,
            'sheet_name': 'Data',
            'column_mappings': [
                {'source': 'Date', 'target': 'Date'},
                {'source': 'SalesAmt', 'target': 'Sales'}
            ],
            'alias': 'sheet0'
        },
        {
            'file_path': costs_file,
            'sheet_name': 'Data',
            'column_mappings': [
                {'source': 'Date', 'target': 'Date'},
                {'source': 'CostAmt', 'target': 'Cost'}
            ],
            'alias': 'sheet1'
        }
    ]

    # 测试内部引用模式（默认，use_external_refs=False）
    print("\n--- 内部引用模式 (use_external_refs=False) ---")
    internal_output = os.path.join(test_dir, 'output_internal.xlsx')
    generate_excel_from_template(
        template_file=template_file,
        template_sheet='Sheet1',
        formula_columns=['Sales', 'Cost', 'Profit'],
        data_sources=data_sources,
        output_file=internal_output,
        string_columns=['Date'],
        use_external_refs=False
    )

    # 验证内部引用模式
    import openpyxl
    wb_internal = openpyxl.load_workbook(internal_output)
    print(f"   Sheet数量: {len(wb_internal.sheetnames)}")
    print(f"   Sheet列表: {wb_internal.sheetnames}")

    # 检查公式是否使用内部引用
    ws_internal = wb_internal['结果']
    sales_formula = ws_internal.cell(row=2, column=2).value
    print(f"   Sales公式: {sales_formula}")
    assert 'sheet0!' in sales_formula, f"内部引用应使用sheet名，而不是文件路径: {sales_formula}"
    assert '[' not in sales_formula, f"内部引用不应包含文件路径: {sales_formula}"
    print("   ✓ 内部引用格式正确")

    # 检查数据源sheet是否存在
    assert 'sheet0' in wb_internal.sheetnames, "数据源sheet0应存在于输出文件中"
    assert 'sheet1' in wb_internal.sheetnames, "数据源sheet1应存在于输出文件中"
    print("   ✓ 数据源sheet已复制到输出文件")

    wb_internal.close()

    # 测试外部引用模式（use_external_refs=True）
    print("\n--- 外部引用模式 (use_external_refs=True) ---")
    external_output = os.path.join(test_dir, 'output_external.xlsx')
    generate_excel_from_template(
        template_file=template_file,
        template_sheet='Sheet1',
        formula_columns=['Sales', 'Cost', 'Profit'],
        data_sources=data_sources,
        output_file=external_output,
        string_columns=['Date'],
        use_external_refs=True
    )

    # 验证外部引用模式
    wb_external = openpyxl.load_workbook(external_output)
    print(f"   Sheet数量: {len(wb_external.sheetnames)}")
    print(f"   Sheet列表: {wb_external.sheetnames}")

    # 检查公式是否使用外部引用
    ws_external = wb_external['结果']
    sales_formula = ws_external.cell(row=2, column=2).value
    print(f"   Sales公式: {sales_formula}")
    assert '[' in sales_formula, f"外部引用应包含文件路径: {sales_formula}"
    print("   ✓ 外部引用格式正确")

    # 检查数据源sheet是否不存在
    assert 'sheet0' not in wb_external.sheetnames, "数据源sheet不应存在于输出文件中（外部引用模式）"
    assert 'sheet1' not in wb_external.sheetnames, "数据源sheet不应存在于输出文件中（外部引用模式）"
    print("   ✓ 数据源sheet未复制到输出文件")

    wb_external.close()

    print("\n内部引用 vs 外部引用测试通过！")


def main():
    print("=== 模板生成器测试 ===\n")

    # 单元测试
    test_parse_column_mappings()
    test_parse_formula_references()
    test_replace_sheet_references()

    # 集成测试
    test_full_generation()

    # 内部引用 vs 外部引用模式测试
    test_internal_vs_external_refs()

    # CLI测试提示
    test_cli()

    print("\n所有测试完成！")


if __name__ == "__main__":
    main()
