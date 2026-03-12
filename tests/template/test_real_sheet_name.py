"""
测试使用实际sheet名（而非别名）的公式
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import openpyxl

TEST_DIR = os.path.dirname(os.path.abspath(__file__))


def create_template_with_real_names():
    """创建使用实际sheet名的模板"""
    print("创建模板（使用实际sheet名）...")

    output_file = os.path.join(TEST_DIR, 'real_name_template.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '加载状态分析'

    # 设置列名
    headers = ['ID', '状态', '产品编码', '类型', '计算结果']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 使用实际的sheet名（不是sheet0、sheet1）
    # ESDP-Bpart 和 ESDP-Cpart 是实际数据文件中的sheet名
    complex_formula = '''=IF(B2="新发货",IF(COUNTIFS('ESDP-Bpart'!A:A,C2,'ESDP-Bpart'!B:B,D2)=0,"无记录",SUMIFS('ESDP-Bpart'!C:C,'ESDP-Bpart'!A:A,C2,'ESDP-Bpart'!B:B,D2)),IF(COUNTIFS('ESDP-Cpart'!A:A,C2,'ESDP-Cpart'!B:B,D2)=0,"无记录",SUMIFS('ESDP-Cpart'!C:C,'ESDP-Cpart'!A:A,C2,'ESDP-Cpart'!B:B,D2)))'''

    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value='新发货')
    ws.cell(row=2, column=3, value='P001')
    ws.cell(row=2, column=4, value='A')
    ws.cell(row=2, column=5, value=complex_formula)

    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value='已发货')
    ws.cell(row=3, column=3, value='P002')
    ws.cell(row=3, column=4, value='B')
    ws.cell(row=3, column=5, value=complex_formula)

    wb.save(output_file)
    print(f"  模板文件: {output_file}")
    print(f"  公式（使用实际sheet名）:\n  {complex_formula[:100]}...\n")
    return output_file


def create_source_files():
    """创建数据源文件"""
    # 数据源A
    data_a = {
        '产品编码': ['P001', 'P001', 'P002', 'P003'],
        '类型': ['A', 'A', 'B', 'C'],
        '数量': [10, 20, 30, 40]
    }
    df_a = pd.DataFrame(data_a)
    file_a = os.path.join(TEST_DIR, 'real_name_source_a.xlsx')
    with pd.ExcelWriter(file_a, engine='openpyxl') as writer:
        df_a.to_excel(writer, sheet_name='ESDP-Bpart', index=False)
    print(f"  数据源A: {file_a} (sheet: ESDP-Bpart)")

    # 数据源B
    data_b = {
        '产品编码': ['P001', 'P002', 'P002'],
        '类型': ['A', 'B', 'B'],
        '数量': [100, 200, 300]
    }
    df_b = pd.DataFrame(data_b)
    file_b = os.path.join(TEST_DIR, 'real_name_source_b.xlsx')
    with pd.ExcelWriter(file_b, engine='openpyxl') as writer:
        df_b.to_excel(writer, sheet_name='ESDP-Cpart', index=False)
    print(f"  数据源B: {file_b} (sheet: ESDP-Cpart)")

    return file_a, file_b


def test_real_sheet_name():
    """测试使用实际sheet名的公式处理"""
    print("=" * 70)
    print("测试：使用实际sheet名（非别名）")
    print("=" * 70 + "\n")

    # 创建测试文件
    template_file = create_template_with_real_names()
    source_a, source_b = create_source_files()

    print("\n运行模板生成器...")

    from modules.template_generator import generate_excel_from_template

    # 注意：这里不需要指定alias，因为模板中使用的是实际sheet名
    data_sources = [
        {
            'file_path': source_a,
            'sheet_name': 'ESDP-Bpart',
            'column_mappings': [
                {'source': '产品编码', 'target': '产品编码'},
                {'source': '类型', 'target': '类型'}
            ]
            # 不需要alias！
        },
        {
            'file_path': source_b,
            'sheet_name': 'ESDP-Cpart',
            'column_mappings': [
                {'source': '产品编码', 'target': '产品编码'},
                {'source': '类型', 'target': '类型'}
            ]
        }
    ]

    output_file = os.path.join(TEST_DIR, 'real_name_output.xlsx')

    result = generate_excel_from_template(
        template_file=template_file,
        template_sheet='加载状态分析',
        formula_columns=['计算结果'],
        data_sources=data_sources,
        output_file=output_file,
        use_external_refs=True
    )

    print("\n" + "=" * 70)
    print("验证结果")
    print("=" * 70)

    # 检查输出
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    original_formula = ws.cell(row=2, column=5).value

    print(f"\n原始公式（模板中）:")
    print("  使用: 'ESDP-Bpart'!A:A, 'ESDP-Cpart'!A:A")
    print(f"\n转换后公式:")
    print(f"  {original_formula[:150]}...")

    # 检查是否正确替换
    if '[real_name_source_a.xlsx]ESDP-Bpart' in original_formula:
        print("\n✅ 成功：'ESDP-Bpart' → '[real_name_source_a.xlsx]ESDP-Bpart'")
    else:
        print("\n❌ 失败：'ESDP-Bpart' 未正确替换")

    if '[real_name_source_b.xlsx]ESDP-Cpart' in original_formula:
        print("✅ 成功：'ESDP-Cpart' → '[real_name_source_b.xlsx]ESDP-Cpart'")
    else:
        print("❌ 失败：'ESDP-Cpart' 未正确替换")

    wb.close()

    print("\n" + "=" * 70)
    print("测试完成！")
    print("=" * 70)


if __name__ == "__main__":
    test_real_sheet_name()
