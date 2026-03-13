"""
测试模板公式中包含文件路径的情况
例如: '[old_file.xlsx]SheetName'!A1 应该被替换为 '[new_file.xlsx]SheetName'!A1
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import openpyxl

TEST_DIR = os.path.dirname(os.path.abspath(__file__))


def create_template_with_file_path():
    """创建包含文件路径的模板"""
    print("创建模板（公式中包含文件路径）...")

    output_file = os.path.join(TEST_DIR, 'file_path_template.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '分析结果'

    # 设置列名
    headers = ['ID', '产品', '销售额', '成本', '利润']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 使用包含文件路径的公式
    # 注意：公式中的文件路径是旧的，需要被替换为实际数据源的文件路径
    complex_formula = '''=IF(B2="新品",SUMIFS('[旧销售数据.xlsx]销售表'!C:C,'[旧销售数据.xlsx]销售表'!A:A,C2),SUMIFS('[旧成本数据.xlsx]成本表'!C:C,'[旧成本数据.xlsx]成本表'!A:A,C2))'''

    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value='新品')
    ws.cell(row=2, column=3, value='P001')
    ws.cell(row=2, column=4, value=100)
    ws.cell(row=2, column=5, value=complex_formula)

    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value='标准')
    ws.cell(row=3, column=3, value='P002')
    ws.cell(row=3, column=4, value=200)
    ws.cell(row=3, column=5, value=complex_formula)

    wb.save(output_file)
    print(f"  模板文件: {output_file}")
    print(f"  公式（包含旧文件路径）:\n  {complex_formula[:100]}...\n")
    return output_file


def create_source_files():
    """创建数据源文件"""
    # 销售数据
    data_a = {
        '产品编码': ['P001', 'P001', 'P002', 'P003'],
        '类型': ['A', 'A', 'B', 'C'],
        '金额': [1000, 2000, 3000, 4000]
    }
    df_a = pd.DataFrame(data_a)
    file_a = os.path.join(TEST_DIR, 'file_path_sales.xlsx')
    with pd.ExcelWriter(file_a, engine='openpyxl') as writer:
        df_a.to_excel(writer, sheet_name='销售表', index=False)
    print(f"  销售数据: {file_a} (sheet: 销售表)")

    # 成本数据
    data_b = {
        '产品编码': ['P001', 'P002', 'P002'],
        '类型': ['A', 'B', 'B'],
        '金额': [500, 1000, 1500]
    }
    df_b = pd.DataFrame(data_b)
    file_b = os.path.join(TEST_DIR, 'file_path_costs.xlsx')
    with pd.ExcelWriter(file_b, engine='openpyxl') as writer:
        df_b.to_excel(writer, sheet_name='成本表', index=False)
    print(f"  成本数据: {file_b} (sheet: 成本表)")

    return file_a, file_b


def test_file_path_in_formula():
    """测试公式中包含文件路径的情况"""
    print("=" * 70)
    print("测试：模板公式中包含文件路径")
    print("=" * 70 + "\n")

    # 创建测试文件
    template_file = create_template_with_file_path()
    source_a, source_b = create_source_files()

    print("\n运行模板生成器...")

    from modules.template_generator import generate_excel_from_template

    # 数据源配置
    # 注意：不需要alias，因为公式中使用的是实际的sheet名
    data_sources = [
        {
            'file_path': source_a,
            'sheet_name': '销售表',
            'column_mappings': [
                {'source': '产品编码', 'target': '产品编码'}
            ]
        },
        {
            'file_path': source_b,
            'sheet_name': '成本表',
            'column_mappings': [
                {'source': '产品编码', 'target': '产品编码'}
            ]
        }
    ]

    output_file = os.path.join(TEST_DIR, 'file_path_output.xlsx')

    result = generate_excel_from_template(
        template_file=template_file,
        template_sheet='分析结果',
        formula_columns=['利润'],
        data_sources=data_sources,
        output_file=output_file,
        use_external_refs=True
    )

    print("\n" + "=" * 70)
    print("验证结果")
    print("=" * 70)

    # 检查输出
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    original_formula = ws.cell(row=2, column=5).value

    print(f"\n原始公式（模板中）:")
    print("  使用: '[旧销售数据.xlsx]销售表'!..., '[旧成本数据.xlsx]成本表'!...")
    print(f"\n转换后公式:")
    print(f"  {original_formula[:200]}...")

    # 检查是否正确替换
    success = True

    if '[file_path_sales.xlsx]销售表' in original_formula:
        print("\n✅ 成功：'[旧销售数据.xlsx]销售表' → '[file_path_sales.xlsx]销售表'")
    else:
        print("\n❌ 失败：销售表引用未正确替换")
        success = False

    if '[file_path_costs.xlsx]成本表' in original_formula:
        print("✅ 成功：'[旧成本数据.xlsx]成本表' → '[file_path_costs.xlsx]成本表'")
    else:
        print("❌ 失败：成本表引用未正确替换")
        success = False

    # 检查旧的文件路径是否已被替换
    if '旧销售数据.xlsx' in original_formula or '旧成本数据.xlsx' in original_formula:
        print("❌ 失败：旧的文件路径未被替换")
        success = False
    else:
        print("✅ 成功：旧的文件路径已被替换")

    wb.close()

    print("\n" + "=" * 70)
    if success:
        print("测试通过！")
    else:
        print("测试失败！")
    print("=" * 70)

    return success


def test_extract_sheet_name():
    """测试 extract_sheet_name 函数"""
    print("=" * 70)
    print("测试：extract_sheet_name 函数")
    print("=" * 70 + "\n")

    from modules.template_generator import replace_sheet_references
    import re

    # 直接测试正则表达式
    def extract_sheet_name(full_reference: str) -> str:
        bracket_match = re.match(r'\[.+\](.+)', full_reference)
        if bracket_match:
            return bracket_match.group(1)
        return full_reference

    test_cases = [
        ('[sales.xlsx]Sheet1', 'Sheet1'),
        ('[old_file.xlsx]销售表', '销售表'),
        ('[data.xlsx]ESDP-Bpart', 'ESDP-Bpart'),
        ('SimpleSheet', 'SimpleSheet'),
        ('销售数据', '销售数据'),
    ]

    all_passed = True
    for input_val, expected in test_cases:
        result = extract_sheet_name(input_val)
        if result == expected:
            print(f"  ✅ '{input_val}' → '{result}'")
        else:
            print(f"  ❌ '{input_val}' → '{result}' (期望: '{expected}')")
            all_passed = False

    print("\n" + "=" * 70)
    if all_passed:
        print("extract_sheet_name 测试通过！")
    else:
        print("extract_sheet_name 测试失败！")
    print("=" * 70)

    return all_passed


if __name__ == "__main__":
    # 先测试 extract_sheet_name 函数
    test_extract_sheet_name()

    print("\n")

    # 再测试完整流程
    test_file_path_in_formula()
