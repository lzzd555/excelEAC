"""
测试 string_columns 参数防止前置零丢失
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import openpyxl

TEST_DIR = os.path.dirname(os.path.abspath(__file__))


def create_source_with_leading_zeros():
    """创建带有前置零的数据源"""
    print("创建数据源（带前置零）...")

    data = {
        '订单号': ['001', '002', '010', '100'],  # 前置零
        '产品编码': ['P001', 'P002', 'P003', 'P004'],
        '数量': [10, 20, 30, 40]
    }
    df = pd.DataFrame(data)
    file_path = os.path.join(TEST_DIR, 'leading_zeros_source.xlsx')
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='数据', index=False)
    print(f"  数据源: {file_path}")
    print(f"  订单号数据: {data['订单号']}")
    return file_path


def create_template_for_string_test():
    """创建测试模板"""
    print("创建模板...")

    output_file = os.path.join(TEST_DIR, 'string_test_template.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '结果'

    headers = ['订单号', '产品编码', '数量', '备注']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    wb.save(output_file)
    print(f"  模板文件: {output_file}")
    return output_file


def test_string_columns():
    """测试 string_columns 参数"""
    print("=" * 70)
    print("测试：string_columns 参数防止前置零丢失")
    print("=" * 70 + "\n")

    # 创建测试文件
    source_file = create_source_with_leading_zeros()
    template_file = create_template_for_string_test()

    print("\n" + "-" * 70)
    print("测试1: 不使用 string_columns 参数（前置零可能丢失）")
    print("-" * 70)

    from modules.template_generator import generate_excel_from_template

    data_sources = [
        {
            'file_path': source_file,
            'sheet_name': '数据',
            'column_mappings': [
                {'source': '订单号', 'target': '订单号'},
                {'source': '产品编码', 'target': '产品编码'},
                {'source': '数量', 'target': '数量'}
            ]
        }
    ]

    output_without_string = os.path.join(TEST_DIR, 'output_without_string.xlsx')

    result = generate_excel_from_template(
        template_file=template_file,
        template_sheet='结果',
        formula_columns=[],
        data_sources=data_sources,
        output_file=output_without_string,
        string_columns=None  # 不指定字符串列
    )

    # 检查结果
    wb = openpyxl.load_workbook(output_without_string)
    ws = wb.active

    print("\n输出数据（无 string_columns）:")
    order_values = []
    for row in range(2, ws.max_row + 1):
        order_val = ws.cell(row=row, column=1).value
        order_values.append(str(order_val) if order_val is not None else '')
        print(f"  订单号: {order_val} (类型: {type(order_val).__name__})")

    wb.close()

    print("\n" + "-" * 70)
    print("测试2: 使用 string_columns 参数（保持前置零）")
    print("-" * 70)

    output_with_string = os.path.join(TEST_DIR, 'output_with_string.xlsx')

    result = generate_excel_from_template(
        template_file=template_file,
        template_sheet='结果',
        formula_columns=[],
        data_sources=data_sources,
        output_file=output_with_string,
        string_columns=['订单号']  # 指定订单号为字符串列
    )

    # 检查结果
    wb = openpyxl.load_workbook(output_with_string)
    ws = wb.active

    print("\n输出数据（有 string_columns）:")
    order_values_with_string = []
    for row in range(2, ws.max_row + 1):
        order_val = ws.cell(row=row, column=1).value
        order_values_with_string.append(str(order_val) if order_val is not None else '')
        print(f"  订单号: {order_val} (类型: {type(order_val).__name__})")

    wb.close()

    print("\n" + "=" * 70)
    print("验证结果")
    print("=" * 70)

    expected_values = ['001', '002', '010', '100']
    success = True

    # 检查是否保持了前置零
    for i, expected in enumerate(expected_values):
        if order_values_with_string[i] == expected:
            print(f"  ✅ 行{i+1}: '{order_values_with_string[i]}' 正确")
        else:
            print(f"  ❌ 行{i+1}: '{order_values_with_string[i]}' != '{expected}'")
            success = False

    print("\n" + "=" * 70)
    if success:
        print("测试通过！前置零保持正常。")
    else:
        print("测试失败！前置零未正确保持。")
    print("=" * 70)

    return success


if __name__ == "__main__":
    test_string_columns()
