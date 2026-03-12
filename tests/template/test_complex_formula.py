"""
测试复杂嵌套公式处理
"""

import sys
import os

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import openpyxl

TEST_DIR = os.path.dirname(os.path.abspath(__file__))


def create_complex_template():
    """创建包含复杂嵌套公式的模板"""
    print("创建复杂公式模板...")

    output_file = os.path.join(TEST_DIR, 'complex_template.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '加载状态分析'

    # 设置列名
    headers = ['ID', '状态', '产品编码', '类型', '数量', '计算结果']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # 第二行设置复杂公式
    # 模拟用户提供的复杂公式格式
    # 注意：这里使用sheet0, sheet1作为别名，对应外部数据文件
    complex_formula = '''=IF(B2="新发货",IF(COUNTIFS(sheet0!A:A,C2,sheet0!B:B,D2)=0,"无记录",SUMIFS(sheet0!C:C,sheet0!A:A,C2,sheet0!B:B,D2)),IF(COUNTIFS(sheet1!A:A,C2,sheet1!B:B,D2)=0,"无记录",SUMIFS(sheet1!C:C,sheet1!A:A,C2,sheet1!B:B,D2)))'''

    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value='新发货')
    ws.cell(row=2, column=3, value='P001')
    ws.cell(row=2, column=4, value='A')
    ws.cell(row=2, column=5, value=100)
    ws.cell(row=2, column=6, value=complex_formula)

    # 添加更多示例行
    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value='已发货')
    ws.cell(row=3, column=3, value='P002')
    ws.cell(row=3, column=4, value='B')
    ws.cell(row=3, column=5, value=200)
    ws.cell(row=3, column=6, value=complex_formula)

    wb.save(output_file)
    print(f"  模板文件: {output_file}")
    print(f"  复杂公式:\n  {complex_formula}\n")
    return output_file


def create_source_data_a():
    """创建数据源A (对应sheet0/ESDP-Bpart)"""
    print("创建数据源A...")

    data = {
        '产品编码': ['P001', 'P001', 'P002', 'P003', 'P001'],
        '类型': ['A', 'A', 'B', 'C', 'A'],
        '数量': [10, 20, 30, 40, 50],
        '状态': ['正常', '正常', '正常', '正常', '异常']
    }
    df = pd.DataFrame(data)
    output_file = os.path.join(TEST_DIR, 'complex_source_a.xlsx')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='ESDP-Bpart', index=False)

    print(f"  文件: {output_file}")
    print(f"  数据:\n{df}\n")
    return output_file


def create_source_data_b():
    """创建数据源B (对应sheet1)"""
    print("创建数据源B...")

    data = {
        '产品编码': ['P001', 'P002', 'P002', 'P003'],
        '类型': ['A', 'B', 'B', 'C'],
        '数量': [100, 200, 300, 400],
        '备注': ['备注1', '备注2', '备注3', '备注4']
    }
    df = pd.DataFrame(data)
    output_file = os.path.join(TEST_DIR, 'complex_source_b.xlsx')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='ESDP-Cpart', index=False)

    print(f"  文件: {output_file}")
    print(f"  数据:\n{df}\n")
    return output_file


def test_complex_formula():
    """测试复杂公式处理"""
    print("=" * 70)
    print("复杂嵌套公式测试")
    print("=" * 70 + "\n")

    # 创建测试文件
    template_file = create_complex_template()
    source_a_file = create_source_data_a()
    source_b_file = create_source_data_b()

    # 使用模板生成器
    from modules.template_generator import generate_excel_from_template

    data_sources = [
        {
            'file_path': source_a_file,
            'sheet_name': 'ESDP-Bpart',
            'column_mappings': [
                {'source': '产品编码', 'target': '产品编码'},
                {'source': '类型', 'target': '类型'},
                {'source': '数量', 'target': '数量'}
            ],
            'alias': 'sheet0'
        },
        {
            'file_path': source_b_file,
            'sheet_name': 'ESDP-Cpart',
            'column_mappings': [
                {'source': '产品编码', 'target': '产品编码'},
                {'source': '类型', 'target': '类型'},
                {'source': '数量', 'target': '数量'}
            ],
            'alias': 'sheet1'
        }
    ]

    output_file = os.path.join(TEST_DIR, 'complex_output.xlsx')

    result = generate_excel_from_template(
        template_file=template_file,
        template_sheet='加载状态分析',
        formula_columns=['计算结果'],
        data_sources=data_sources,
        output_file=output_file,
        use_external_refs=True
    )

    print("\n" + "=" * 70)
    print("验证输出结果")
    print("=" * 70)

    # 检查输出文件中的公式
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    print("\n原始公式:")
    print(ws.cell(row=2, column=6).value)

    print("\n" + "-" * 70)

    # 分析公式中的引用
    import re
    formula = ws.cell(row=2, column=6).value

    # 检查sheet引用
    sheet_refs = re.findall(r"sheet\d+![A-Z]+(?::[A-Z]+)?", str(formula), re.IGNORECASE)
    print(f"\n找到的sheet引用: {sheet_refs}")

    # 检查整列引用 (如 A:A)
    column_refs = re.findall(r"[A-Z]+:[A-Z]+", str(formula))
    print(f"找到的整列引用: {column_refs}")

    wb.close()

    print("\n" + "=" * 70)
    print("测试完成！")
    print("=" * 70)

    return output_file


if __name__ == "__main__":
    test_complex_formula()
