#!/usr/bin/env python3
"""
调试测试：深入分析 string_columns 为什么没有正确工作
"""

import pandas as pd
from excel_validator import process_excel_with_validation

import sys
import os

# 添加父目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.validation import
print("=== 调试 string_columns 问题 ===\n")

# 创建测试数据
data = {
    '订单号': ['001', '002', '003', '004'],  # 确保是字符串
    '产品代码': ['001', '002', '001', '003'],  # 确保是字符串
    '部门': ['A', 'A', 'B', 'B'],
    '计划值': [100, 200, 150, 120],
    '实际值': [100, 210, 150, 120]  # 第2行有异常
}

df = pd.DataFrame(data)

print("原始数据：")
print(df.dtypes)
print("\n原始数据内容：")
print(df)

# 保存并检查
df.to_excel('debug_test.xlsx', index=False)

print("\n运行验证...")
result = process_excel_with_validation(
    input_file='debug_test.xlsx',
    sheet_name='Sheet1',
    group_columns=['部门'],
    compare_columns=['计划值', '实际值'],
    output_columns=['部门', '验证状态', '总行数'],
    output_file='debug_output.xlsx',
    string_columns=['订单号', '产品代码']
)

print("\n分组结果：")
print(result)

# 直接检查Excel文件内容
print("\n直接读取Excel文件内容：")
import openpyxl
wb = openpyxl.load_workbook('debug_output.xlsx')

# 检查异常详情
if '异常详情' in wb.sheetnames:
    detail_sheet = wb['异常详情']
    print("\n异常详情内容：")
    data_rows = []
    for i, row in enumerate(detail_sheet.iter_rows(values_only=True)):
        if i == 0:  # 标题行
            print(f"列名: {row}")
        else:  # 数据行
            data_rows.append(row)
            print(f"第{i}行: {row}")

    # 详细检查订单号列
    print("\n详细检查订单号：")
    order_id_col = None
    for i, col_name in enumerate(detail_sheet[1]):  # 第一行是列名
        if col_name == '订单号':
            order_id_col = i
            break

    if order_id_col is not None:
        print(f"订单号列索引: {order_id_col}")
        for i, row in enumerate(detail_sheet.iter_rows(min_row=2, values_only=True)):
            if row:  # 非空行
                order_id = row[order_id_col]
                print(f"订单号值: {order_id}, 类型: {type(order_id)}")

# 清理
import os
os.remove('debug_test.xlsx')
os.remove('debug_output.xlsx')