"""
综合测试：同时包含自引用和外部引用
"""
import pandas as pd
import openpyxl
import os

# 清理旧文件
for f in ['external_data.xlsx', 'template_mixed_refs.xlsx', 'test_mixed_refs_output.xlsx']:
    if os.path.exists(f):
        os.remove(f)

# 创建外部数据源文件
def create_external_data_file():
    """创建外部数据源文件 external_data.xlsx"""
    data = {
        'Sales': [100, 200, 300, 400, 500],
        'Cost': [50, 100, 150, 200, 250],
        'Tax': [10, 20, 30, 40, 50],
        'Target': [20, 40, 60, 80, 100]  # 外部数据，用于外部引用测试
    }
    df = pd.DataFrame(data)

    # 保存为 Excel 文件
    with pd.ExcelWriter('external_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='DataSheet', index=False)

    print("✓ 创建外部数据文件: external_data.xlsx")


# 创建包含混合引用的模板文件
def create_template_with_mixed_refs():
    """创建模板文件，同时包含自引用和外部引用"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MyTemplateSheet'  # 设置模板 sheet 名称

    # 第一行：列名
    headers = ['Sales', 'Cost', 'Tax', 'Target', 'SubTotal', 'Profit', 'Ratio', 'Variance']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # 第二行：公式模板（混合引用）

    # SubTotal: 自引用 =Sales-Cost-Tax
    ws.cell(row=2, column=5).value = "=A2-B2-C2"

    # Profit: 自引用 =SubTotal*0.2
    ws.cell(row=2, column=6).value = "=E2*0.2"

    # Ratio: 自引用 =Profit/SubTotal
    ws.cell(row=2, column=7).value = "=F2/E2"

    # Variance: 外部引用 =SubTotal-[external_data.xlsx]DataSheet!D2
    # 注意：这里需要创建一个外部链接来模拟这个场景
    # 为了简化测试，我们使用别名的形式：[1]DataSheet!D2
    ws.cell(row=2, column=8).value = "=[1]DataSheet!D2"

    # 保存文件
    wb.save('template_mixed_refs.xlsx')
    wb.close()

    print("✓ 创建模板文件: template_mixed_refs.xlsx")
    print("  - Sheet 名称: MyTemplateSheet")
    print("  - 包含混合引用:")
    print("    - SubTotal: 自引用 (Sales-Cost-Tax)")
    print("    - Profit: 自引用 (SubTotal*0.2)")
    print("    - Ratio: 自引用 (Profit/SubTotal)")
    print("    - Variance: 外部引用 ([1]DataSheet!D2)")


def run_test():
    """运行测试"""
    from modules.template_generator import generate_excel_from_template

    print("\n" + "="*70)
    print("开始综合测试：混合引用（自引用 + 外部引用）")
    print("="*70 + "\n")

    # 数据源配置
    data_sources = [
        {
            'file_path': 'external_data.xlsx',
            'sheet_name': 'DataSheet',
            'column_mappings': [
                {'source': 'Sales', 'target': 'Sales'},
                {'source': 'Cost', 'target': 'Cost'},
                {'source': 'Tax', 'target': 'Tax'},
                {'source': 'Target', 'target': 'Target'}
            ],
            'alias': 'external_data'
        }
    ]

    # 公式列（这些列的公式会被应用到输出文件）
    formula_columns = ['SubTotal', 'Profit', 'Ratio', 'Variance']

    try:
        # 生成 Excel 文件
        result = generate_excel_from_template(
            template_file='template_mixed_refs.xlsx',
            template_sheet='MyTemplateSheet',
            formula_columns=formula_columns,
            data_sources=data_sources,
            output_file='test_mixed_refs_output.xlsx',
            use_external_refs=True  # 使用外部引用模式
        )

        print("\n" + "="*70)
        print("✓ 测试完成！")
        print("="*70)
        print("\n输出文件: test_mixed_refs_output.xlsx")

        # 验证输出文件中的公式
        print("\n正在验证输出文件中的公式...")
        wb = openpyxl.load_workbook('test_mixed_refs_output.xlsx', data_only=False)
        ws = wb.active

        # 读取第二行的公式
        subtotal_formula = ws.cell(row=2, column=5).value
        profit_formula = ws.cell(row=2, column=6).value
        ratio_formula = ws.cell(row=2, column=7).value
        variance_formula = ws.cell(row=2, column=8).value

        print(f"\n公式验证结果:")
        print(f"  SubTotal (自引用): {subtotal_formula}")
        print(f"  Profit (自引用):   {profit_formula}")
        print(f"  Ratio (自引用):    {ratio_formula}")
        print(f"  Variance (外部引用): {variance_formula}")

        wb.close()

        # 检查结果
        print("\n结果分析:")

        # 检查自引用是否使用本地引用
        self_refs_ok = True
        if 'test_mixed_refs_output.xlsx' in str(subtotal_formula):
            print("  ✗ SubTotal 包含文件名（应该是本地引用）")
            self_refs_ok = False
        if 'test_mixed_refs_output.xlsx' in str(profit_formula):
            print("  ✗ Profit 包含文件名（应该是本地引用）")
            self_refs_ok = False
        if 'test_mixed_refs_output.xlsx' in str(ratio_formula):
            print("  ✗ Ratio 包含文件名（应该是本地引用）")
            self_refs_ok = False

        if self_refs_ok:
            print("  ✓ 自引用公式使用本地引用格式")

        # 检查外部引用是否使用文件名
        external_ref_ok = True
        if 'external_data.xlsx' not in str(variance_formula):
            print("  ✗ Variance 不包含外部文件名")
            external_ref_ok = False
        else:
            print("  ✓ 外部引用公式使用文件名格式")

        if self_refs_ok and external_ref_ok:
            print("\n✓ 所有验证通过！混合引用处理正确。")
        else:
            print("\n✗ 验证失败，请检查公式格式。")

    except Exception as e:
        print(f"\n✗ 测试失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 创建测试文件
    create_external_data_file()
    create_template_with_mixed_refs()

    # 运行测试
    run_test()
